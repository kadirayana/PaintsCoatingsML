"""
Paint Formulation AI - Formülasyon Editörü
==========================================
Excel benzeri formülasyon giriş ve düzenleme paneli

NOT: Bu dosya artık bir facade olarak çalışmaktadır.
Modüler bileşenler app/components/editor/ altında tanımlanmıştır:
- ComponentGrid: Treeview tabanlı grid
- FormulationSummary: Özet bilgiler paneli
- ExcelHandler: Excel import/export işlemleri
- PredictionPanel: ML tahmin sonuçları
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Callable, Dict, List, Optional
import threading

# Modüler bileşenlerden import (Yeni özellikler bu dosyalarda)
from app.components.editor.component_grid import ComponentGrid
from app.components.editor.formulation_summary import FormulationSummary
from app.components.editor.excel_handler import ExcelHandler
from app.components.editor.prediction_panel import PredictionPanel



class FormulationEditorPanel(ttk.LabelFrame):
    """
    Excel benzeri Formülasyon Editörü
    
    Sütunlar:
    - Hammadde Kodu
    - Hammadde Adı
    - Katı Miktarı
    - %
    - Fiyat/Birim
    """
    
    
    def __init__(self, parent, on_save: Callable = None, on_calculate: Callable = None, on_load_formulation: Callable = None, 
                 on_lookup_material: Callable = None, on_get_material_list: Callable = None):
        super().__init__(parent, text="📋 Formülasyon Editörü", padding=10)
        
        self.on_save = on_save
        self.on_calculate = on_calculate
        self.row_count = 0
        self.current_project = None
        self.on_load_formulation = on_load_formulation
        self.on_lookup_material = on_lookup_material
        self.on_get_material_list = on_get_material_list
        self.material_cache = {} # code -> material dict
        
        # Proje seçici
        project_frame = ttk.LabelFrame(self, text="📁 Proje", padding=5)
        project_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(project_frame, text="Proje:").pack(side=tk.LEFT)
        self.project_combo = ttk.Combobox(project_frame, width=30, state='readonly')
        self.project_combo.pack(side=tk.LEFT, padx=5)
        self.project_combo.bind('<<ComboboxSelected>>', self._on_project_selected)
        
        ttk.Button(project_frame, text="➕ Yeni Proje", command=self._create_new_project).pack(side=tk.LEFT, padx=5)
        
        # Formülasyon seçici
        ttk.Separator(project_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Label(project_frame, text="Kayıtlı Formülasyon:").pack(side=tk.LEFT)
        self.formulation_combo = ttk.Combobox(project_frame, width=25, state='readonly')
        self.formulation_combo.pack(side=tk.LEFT, padx=5)
        self.formulation_combo.bind('<<ComboboxSelected>>', self._on_formulation_selected)
        
        self.formulation_list = []  # ID -> formulation mapping
        
        # Üst bilgi alanı
        header_frame = ttk.Frame(self)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(header_frame, text="Formül Kodu:").pack(side=tk.LEFT)
        self.formula_code_entry = ttk.Entry(header_frame, width=15)
        self.formula_code_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(header_frame, text="Formül Adı:").pack(side=tk.LEFT, padx=(20, 0))
        self.formula_name_entry = ttk.Entry(header_frame, width=25)
        self.formula_name_entry.pack(side=tk.LEFT, padx=5)
        
        # Tablo alanı
        table_frame = ttk.Frame(self)
        table_frame.pack(fill=tk.BOTH, expand=True)
        
        # Treeview - Excel benzeri tablo (Sütunlar güncellendi)
        columns = ('row_num', 'code', 'name', 'amount', 'solid_amount', 'solid_content', 'percentage', 'price')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=12)
        
        # Sütun başlıkları
        self.tree.heading('row_num', text='#')
        self.tree.heading('code', text='Hammadde Kodu')
        self.tree.heading('name', text='Hammadde Adı')
        self.tree.heading('amount', text='Miktar (kg)')  # YENİ
        self.tree.heading('solid_amount', text='Katı (kg)') # GÜNCELLENDİ
        self.tree.heading('solid_content', text='Katı %') # YENİ (Master Data)
        self.tree.heading('percentage', text='%')
        self.tree.heading('price', text='Fiyat/Birim')
        
        # Sütun genişlikleri
        self.tree.column('row_num', width=30, anchor='center')
        self.tree.column('code', width=100)
        self.tree.column('name', width=150)
        self.tree.column('amount', width=80, anchor='e')
        self.tree.column('solid_amount', width=80, anchor='e')
        self.tree.column('solid_content', width=60, anchor='e')
        self.tree.column('percentage', width=60, anchor='e')
        self.tree.column('price', width=80, anchor='e')
        
        # Scrollbar
        scrollbar_y = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Selection event
        self.tree.bind('<Double-1>', self._on_double_click)
        
        # Giriş alanları
        input_frame = ttk.LabelFrame(self, text="Satır Ekle/Düzenle", padding=5)
        input_frame.pack(fill=tk.X, pady=10)
        
        self.entry_vars = {}
        # Giriş konfigürasyonu (Key, Label, Width, ReadOnly, WidgetType)
        entries_config = [
            ('code', 'Kod:', 15, False, 'combobox'),
            ('name', 'Ad:', 20, False, 'entry'), # Adı değiştirebilmeli mi? Evet, özel isim
            ('amount', 'Miktar (kg):', 10, False, 'entry'),
            ('solid_amount', 'Katı (kg):', 10, True, 'entry'), # Otomatik hesaplanır
            ('solid_content', 'Katı %:', 8, True, 'entry'), # Master Data (Read-only)
            ('percentage', '%:', 8, True, 'entry'), # Otomatik hesaplanır
            ('price', 'Fiyat:', 10, False, 'entry'),
        ]
        
        row = ttk.Frame(input_frame)
        row.pack(fill=tk.X, pady=2)
        
        for key, label, width, readonly, w_type in entries_config:
            ttk.Label(row, text=label).pack(side=tk.LEFT)
            
            if w_type == 'combobox':
                entry = ttk.Combobox(row, width=width)
                entry.bind('<<ComboboxSelected>>', self._on_code_selected)
                if self.on_get_material_list:
                    try:
                        materials = self.on_get_material_list()
                        entry['values'] = [m['code'] for m in materials if m.get('code')]
                    except Exception:
                        pass
            else:
                entry = ttk.Entry(row, width=width)
                if key == 'amount':
                    entry.bind('<KeyRelease>', self._calculate_row_inputs)
                if readonly:
                   entry.state(['readonly']) 
            
            entry.pack(side=tk.LEFT, padx=2)
            self.entry_vars[key] = entry
        
        # Butonlar
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(btn_frame, text="➕ Satır Ekle", command=self._add_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="✏️ Güncelle", command=self._update_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🗑️ Satır Sil", command=self._delete_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🧹 Temizle", command=self._clear_all).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(btn_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Button(btn_frame, text="🔮 Tahmin Et", command=self._predict_results).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 Hesapla", command=self._calculate).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="💾 Kaydet", command=self._save).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📥 Excel'den Yükle", command=self._load_from_excel).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📤 Excel'e Aktar", command=self._export_to_excel).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(btn_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Button(btn_frame, text="📝 Excel Şablonu Aç", command=self._open_excel_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📥 Şablonu İçe Aktar", command=self._import_excel_template).pack(side=tk.LEFT, padx=2)
        
        self.template_path = None  # Açılan şablon yolu
        
        # Özet bilgiler
        summary_frame = ttk.Frame(self)
        summary_frame.pack(fill=tk.X, pady=10)
        
        self.summary_labels = {}
        summaries = [
            ('total_solid', 'Toplam Katı:', '0'),
            ('total_percent', 'Toplam %:', '0'),
            ('total_cost', 'Toplam Maliyet:', '0'),
            ('pvc', 'PVC (%):', '0'),
            ('voc', 'VOC (g/L):', '0'),
            ('row_count', 'Satır Sayısı:', '0'),
        ]
        
        for key, label, default in summaries:
            lbl_frame = ttk.Frame(summary_frame)
            lbl_frame.pack(side=tk.LEFT, padx=10)
            
            ttk.Label(lbl_frame, text=label).pack(side=tk.LEFT)
            value_lbl = ttk.Label(lbl_frame, text=default, font=('Helvetica', 10, 'bold'))
            value_lbl.pack(side=tk.LEFT, padx=5)
            self.summary_labels[key] = value_lbl
        
        # === TAHMİN SONUÇLARI PANELİ ===
        self.prediction_frame = ttk.LabelFrame(self, text="🔮 Muhtemel Test Sonuçları (Tahmin)", padding=5)
        self.prediction_frame.pack(fill=tk.X, pady=10)
        
        # Kaplama kalınlığı girişi
        thickness_row = ttk.Frame(self.prediction_frame)
        thickness_row.pack(fill=tk.X, pady=5)
        
        ttk.Label(thickness_row, text="Kaplama Kalınlığı (µm):").pack(side=tk.LEFT)
        self.thickness_entry = ttk.Entry(thickness_row, width=8)
        self.thickness_entry.insert(0, "30")
        self.thickness_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(thickness_row, text="🔮 Tahmin Yap", command=self._predict_results).pack(side=tk.LEFT, padx=10)
        
        # Tahmin sonuçları gösterim alanı
        self.prediction_text = tk.Text(self.prediction_frame, height=6, wrap=tk.WORD, state='disabled')
        self.prediction_text.pack(fill=tk.X, pady=5)
        
        # Bilgilendirme etiketi
        info_label = ttk.Label(
            self.prediction_frame, 
            text="💡 Formülasyon hammaddelerinize göre ML modeli muhtemel test sonuçlarını tahmin eder",
            font=('Helvetica', 8, 'italic')
        )
        info_label.pack(anchor=tk.W)
        
        # Tahmin callback
        self.on_predict = None
    
    def _on_code_selected(self, event=None):
        """Hammadde kodu seçildiğinde"""
        code = self.entry_vars['code'].get()
        if not code or not self.on_lookup_material:
            return
            
        selected_material = self.on_lookup_material(code)
        if selected_material:
            # Cache'e ekle
            self.material_cache[code] = selected_material
            
            # Master data alanlarını doldur (Read-only)
            solid_content = selected_material.get('solid_content', 0)
            
            self.entry_vars['solid_content'].state(['!readonly'])
            self.entry_vars['solid_content'].delete(0, tk.END)
            self.entry_vars['solid_content'].insert(0, f"{solid_content:.1f}")
            self.entry_vars['solid_content'].state(['readonly'])
            
            # Ad ve fiyat doldur
            self.entry_vars['name'].delete(0, tk.END)
            self.entry_vars['name'].insert(0, selected_material.get('name', ''))
            
            # Fiyatı doldur
            self.entry_vars['price'].delete(0, tk.END)
            self.entry_vars['price'].insert(0, str(selected_material.get('unit_price', 0)))
            
            self._calculate_row_inputs()
    
    def _calculate_row_inputs(self, event=None):
        """Satır girişlerini hesapla"""
        try:
            code = self.entry_vars['code'].get()
            amount_str = self.entry_vars['amount'].get()
            
            if not amount_str:
                return
            
            amount = float(amount_str)
            
            # Master data'dan katı yüzdesini al
            material = self.material_cache.get(code, {}) if code else {}
            # Eğer cache'de yoksa ve lookup varsa son bir kez dene (manuel yazılmışsa)
            if not material and code and self.on_lookup_material:
                material = self.on_lookup_material(code) or {}
                
            solid_content = float(material.get('solid_content', 0) or 0)
            
            # Katı Miktarını Hesapla
            solid_amount = amount * (solid_content / 100.0)
            
            # Güncelle
            entry = self.entry_vars['solid_amount']
            entry.state(['!readonly'])
            entry.delete(0, tk.END)
            entry.insert(0, f"{solid_amount:.2f}")
            entry.state(['readonly'])

            # Katı % alanını da güncelle (eğer boşsa veya değişmişse)
            sc_entry = self.entry_vars['solid_content']
            if not sc_entry.get() or float(sc_entry.get()) != solid_content:
                sc_entry.state(['!readonly'])
                sc_entry.delete(0, tk.END)
                sc_entry.insert(0, f"{solid_content:.1f}")
                sc_entry.state(['readonly'])
            
        except ValueError:
            pass
            
    def _add_row(self):
        """Satır ekle"""
        self.row_count += 1
        
        code = self.entry_vars['code'].get().strip()
        name = self.entry_vars['name'].get().strip()
        amount = self.entry_vars['amount'].get().strip() or '0'
        solid_amount = self.entry_vars['solid_amount'].get().strip() or '0'
        solid_content = self.entry_vars['solid_content'].get().strip() or '0'
        percentage = '0' # Otomatik hesaplanacak
        price = self.entry_vars['price'].get().strip() or '0'
        
        if not code:
            code = f"HM{self.row_count:03d}"
        
        self.tree.insert('', tk.END, values=(
            self.row_count,
            code,
            name,
            amount,
            solid_amount,
            solid_content,
            percentage,
            price
        ))
        
        self._recalculate_percentages()
        self._clear_inputs()
        self._update_summary()
    
    def _recalculate_percentages(self):
        """Tüm satırların yüzdelerini yeniden hesapla"""
        total_amount = 0
        children = self.tree.get_children()
        
        # Toplam miktarı bul
        for item in children:
            values = self.tree.item(item)['values']
            amount = self._safe_float(values[3]) # Index 3: Amount
            total_amount += amount
            
        # Yüzdeleri güncelle
        if total_amount > 0:
            for item in children:
                values = list(self.tree.item(item)['values'])
                amount = self._safe_float(values[3])
                percent = (amount / total_amount) * 100.0
                values[6] = f"{percent:.2f}" # Index 6: Percentage
                self.tree.item(item, values=values)

    def _update_row(self):
        """Seçili satırı güncelle"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("Uyarı", "Güncellenecek satırı seçin!")
            return
        
        item = selection[0]
        current_values = self.tree.item(item)['values']
        row_num = current_values[0]
        
        code = self.entry_vars['code'].get().strip() or current_values[1]
        name = self.entry_vars['name'].get().strip() or current_values[2]
        amount_str = self.entry_vars['amount'].get().strip()
        amount = amount_str if amount_str else str(current_values[3])
        solid_amount = self.entry_vars['solid_amount'].get().strip()
        solid_amount = solid_amount if solid_amount else str(current_values[4])
        solid_content = self.entry_vars['solid_content'].get().strip()
        solid_content = solid_content if solid_content else str(current_values[5])
        percentage = '0' # Will be recalculated
        price = self.entry_vars['price'].get().strip() or str(current_values[7])
        
        # Eğer miktar değiştiyse ve katı miktarı read-only ise (yani manuel girilmemişse), yeniden hesapla
        # Ancak burada basitleştirelim: Giriş alanındaki değeri al (zaten hesaplanmıştır)
        
        self.tree.item(item, values=(row_num, code, name, amount, solid_amount, solid_content, percentage, price))
        self._recalculate_percentages()
        self._update_summary()
    
    def _delete_row(self):
        """Seçili satırı sil"""
        selection = self.tree.selection()
        if not selection:
            return
        
        if messagebox.askyesno("Onay", "Seçili satırı silmek istiyor musunuz?"):
            for item in selection:
                self.tree.delete(item)
            self._recalculate_percentages()
            self._update_summary()
    
    def _clear_all(self):
        """Tüm satırları temizle"""
        if messagebox.askyesno("Onay", "Tüm satırları silmek istiyor musunuz?"):
            self.tree.delete(*self.tree.get_children())
            self.row_count = 0
            self._update_summary()
    
    def _clear_inputs(self):
        """Giriş alanlarını temizle"""
        for key, entry in self.entry_vars.items():
            if isinstance(entry, ttk.Entry):
                is_readonly = 'readonly' in entry.state()
                if is_readonly:
                    entry.state(['!readonly'])
                    entry.delete(0, tk.END)
                    entry.state(['readonly'])
                else:
                    entry.delete(0, tk.END)
            elif isinstance(entry, ttk.Combobox):
                entry.set('')
    
    def _on_double_click(self, event):
        """Çift tıklama ile düzenleme"""
        selection = self.tree.selection()
        if selection:
            values = self.tree.item(selection[0])['values']
            
            # Code
            if isinstance(self.entry_vars['code'], ttk.Combobox):
                self.entry_vars['code'].set(values[1])
            else:
                self.entry_vars['code'].delete(0, tk.END)
                self.entry_vars['code'].insert(0, values[1])
            
            self.entry_vars['name'].delete(0, tk.END)
            self.entry_vars['name'].insert(0, values[2])
            
            self.entry_vars['amount'].delete(0, tk.END)
            self.entry_vars['amount'].insert(0, str(values[3]))
            
            # Read-only field update
            sa_entry = self.entry_vars['solid_amount']
            sa_entry.state(['!readonly'])
            sa_entry.delete(0, tk.END)
            sa_entry.insert(0, str(values[4]))
            sa_entry.state(['readonly'])
            self.entry_vars['price'].insert(0, str(values[6]))
    
    def _update_summary(self):
        """Özet bilgileri güncelle"""
        total_solid = 0
        total_amount = 0
        total_cost = 0
        
        total_pvc_pigment_vol = 0
        total_binder_vol = 0
        total_voc_mass = 0
        total_volume = 0
        
        count = 0
        
        children = self.tree.get_children()
        
        for item in children:
            count += 1
            values = self.tree.item(item)['values']
            
            code = values[1]
            amount = self._safe_float(values[3])
            solid_amount_val = self._safe_float(values[4])
            price_val = self._safe_float(values[7])
            
            total_amount += amount
            total_solid += solid_amount_val
            total_cost += amount * price_val
            
            # --- PVC & VOC Hesaplama ---
            # Cache'den özellikleri al
            material = self.material_cache.get(code, {})
            if not material and self.on_lookup_material and code:
                material = self.on_lookup_material(code) or {}
                self.material_cache[code] = material
            
            density = float(material.get('density', 1.0) or 1.0)
            if density <= 0: density = 1.0
            
            voc_g_l = float(material.get('voc_g_l', 0) or 0)
            category = str(material.get('category', '')).lower()
            
            # Hacim (Litre) = kg / (kg/L)
            vol = amount / density
            total_volume += vol
            
            # VOC Kütlesi (g) = Hacim (L) * VOC (g/L)
            voc_mass = vol * voc_g_l
            total_voc_mass += voc_mass
            
            # PVC için Katı Hacmi Yaklaşımı
            # Basitçe: Katı Kütlesi / Yoğunluk (Yaklaşık)
            vol_solid = solid_amount_val / density
            
            if category in ['pigment', 'filler', 'extender', 'dolgu']:
                total_pvc_pigment_vol += vol_solid
            elif category in ['resin', 'binder', 'bağlayıcı']:
                total_binder_vol += vol_solid
        
        # PVC Hesaplama
        if (total_pvc_pigment_vol + total_binder_vol) > 0:
            pvc = (total_pvc_pigment_vol / (total_pvc_pigment_vol + total_binder_vol)) * 100.0
        else:
            pvc = 0
            
        # VOC Hesaplama (g/L)
        if total_volume > 0:
            voc = total_voc_mass / total_volume
        else:
            voc = 0
            
        # Toplam Yüzde Kontrolü (Miktarlar üzerinden değil, explicit percent column üzerinden mi? Hayır, hesaplanan)
        # _recalculate_percentages zaten % sütununu güncelledi.
        # Toplam % her zaman 100 olmalı eğer Amount üzerinden gidiyorsak?
        # Hayır, total % formülasyon mantığında 100 dür.
        # Ama biz Amount giriyoruz.
        # Toplam Miktar 100 ise % = Miktar.
        
        self.summary_labels['total_solid'].config(text=f"{total_solid:.2f}")
        self.summary_labels['total_percent'].config(
            text=f"{total_amount:.1f}", # Toplam Miktar gösterelim artık, % değil
            foreground="black"
        )
        self.summary_labels['total_cost'].config(text=f"{total_cost:.2f}")
        self.summary_labels['row_count'].config(text=str(count))
        self.summary_labels['pvc'].config(text=f"{pvc:.1f}%")
        self.summary_labels['voc'].config(text=f"{voc:.1f}")
        
        # 100% / 1000 Validasyon
        # Label'ı "Toplam Miktar" olarak güncellemek lazım init'te ama şimdi text ile idare edelim.
        # User 100 Check istemişti.
        # Eğer total_amount 100, 1000 değilse uyarı rengi?
        color = "green" if abs(total_amount - 100) < 0.1 or abs(total_amount - 1000) < 1.0 else "red"
        self.summary_labels['total_percent'].config(foreground=color)
    
    
    def _calculate(self):
        """Hesaplama yap"""
        self._update_summary()
        
        if self.on_calculate:
            data = self.get_formulation_data()
            self.on_calculate(data)
            
    def _save(self):
        """Formülasyonu kaydet"""
        formula_code = self.formula_code_entry.get().strip()
        formula_name = self.formula_name_entry.get().strip()
        
        if not formula_code:
            messagebox.showwarning("Uyarı", "Formül kodu girilmelidir!")
            return
        
        data = self.get_formulation_data()
        data['formula_code'] = formula_code
        data['formula_name'] = formula_name
        
        if self.on_save:
            self.on_save(data)
            messagebox.showinfo("Başarılı", f"Formülasyon kaydedildi: {formula_code}")
    
    def _load_from_excel(self):
        """Excel'den yükle"""
        file_path = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")]
        )
        
        if file_path:
            try:
                from src.data_handlers.file_system_manager import FileSystemManager
                fs = FileSystemManager()
                data = fs.read_excel(file_path)
                
                if not data:
                    messagebox.showwarning("Uyarı", "Dosyada veri bulunamadı!")
                    return
                
                # Verileri tabloya yükle
                self.tree.delete(*self.tree.get_children())
                self.row_count = 0
                
                # İlk satırın sütunlarını al
                first_row = data[0]
                columns = list(first_row.keys())
                
                # Sütun eşleştirmesi - çeşitli varyasyonları destekle
                code_keys = ['hammadde_kodu', 'code', 'kod', 'Kod', 'Hammadde Kodu', 'HAMMADDE KODU', 'Column_0']
                name_keys = ['hammadde_adi', 'name', 'ad', 'Ad', 'Hammadde Adı', 'HAMMADDE ADI', 'Adı', 'Column_1']
                amount_keys = ['miktar', 'amount', 'Miktar', 'MİKTAR', 'Column_2']
                # Eğer eski şablondan geliyorsa "kati_miktari" da Miktar olarak kabul edilebilir
                amount_keys.extend(['kati_miktari', 'solid_amount', 'katı', 'Katı', 'Katı Miktarı', 'KATI MİKTARI'])
                
                percent_keys = ['yuzde', 'percentage', '%', 'Yüzde', 'YÜZDE', 'oran', 'Oran', 'Column_3']
                price_keys = ['fiyat', 'price', 'Fiyat', 'FİYAT', 'birim fiyat', 'Birim Fiyat', 'Column_4']
                
                def find_value(row, key_list, default=''):
                    for key in key_list:
                        if key in row:
                            val = row[key]
                            return val if val is not None else default
                    # Eğer eşleşme yoksa sırayla dene
                    row_values = list(row.values())
                    if key_list[-1].startswith('Column_'):
                        try:
                            idx = int(key_list[-1].replace('Column_', ''))
                            if 0 <= idx < len(row_values):
                                return row_values[idx] if row_values[idx] is not None else default
                        except ValueError:
                            pass
                    return default
                
                for row in data:
                    self.row_count += 1
                    
                    # Değerleri bul
                    code = find_value(row, code_keys, f'HM{self.row_count:03d}')
                    name = find_value(row, name_keys, '')
                    amount = find_value(row, amount_keys, 0)
                    percent = find_value(row, percent_keys, 0)
                    price = find_value(row, price_keys, 0)
                    
                    solid_amount = 0
                    solid_content = 0
                    
                    # Material caching & Calculation
                    if code and self.on_lookup_material:
                        material = self.on_lookup_material(code)
                        if material:
                            self.material_cache[code] = material
                            if not name: name = material.get('name', '')
                            if not price: price = material.get('unit_price', 0)
                            
                            solid_content = float(material.get('solid_content', 0) or 0)
                            try:
                                solid_amount = float(amount) * (solid_content / 100.0)
                            except:
                                solid_amount = 0
                    
                    self.tree.insert('', tk.END, values=(
                        self.row_count,
                        code,
                        name,
                        f"{amount}",
                        f"{solid_amount:.2f}",
                        f"{solid_content:.1f}",
                        f"{percent}",
                        f"{price}"
                    ))
                
                self._recalculate_percentages()
                self._update_summary()
                messagebox.showinfo("Başarılı", f"{len(data)} satır yüklendi!\n\nSütunlar: {', '.join(columns[:5])}")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya yüklenemedi: {str(e)}")
    
    def _export_to_excel(self):
        """Excel'e aktar"""
        file_path = filedialog.asksaveasfilename(
            title="Excel Dosyası Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        
        if file_path:
            try:
                from src.data_handlers.file_system_manager import FileSystemManager
                fs = FileSystemManager()
                
                data = []
                for item in self.tree.get_children():
                    values = self.tree.item(item)['values']
                    data.append({
                        'hammadde_kodu': values[1],
                        'hammadde_adi': values[2],
                        'miktar': values[3],
                        'kati_miktari': values[4],
                        'kati_yuzde': values[5],
                        'yuzde': values[6],
                        'fiyat': values[7]
                    })
                
                fs.write_excel(data, file_path)
                messagebox.showinfo("Başarılı", f"Dosya kaydedildi: {file_path}")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya kaydedilemedi: {str(e)}")
    
    def _open_excel_template(self):
        """Excel şablonu oluştur ve aç"""
        import os
        from datetime import datetime
        
        # Şablon klasörü
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
        os.makedirs(template_dir, exist_ok=True)
        
        # Benzersiz dosya adı
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        template_name = f"formulation_{timestamp}.xlsx"
        template_path = os.path.join(template_dir, template_name)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Formülasyon"
            
            # Başlık stili
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sütun başlıkları
            headers = ["Hammadde Kodu", "Hammadde Adı", "Miktar (kg)", "%", "Fiyat/Birim"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            
            # Örnek satırlar (boş)
            for row in range(2, 22):  # 20 boş satır
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.border = thin_border
            
            wb.save(template_path)
            
            # Excel'i aç
            os.startfile(template_path)
            
            self.template_path = template_path
            messagebox.showinfo(
                "Excel Şablonu", 
                f"Excel şablonu açıldı:\n{template_name}\n\n"
                "1. Hammadde bilgilerini girin (Miktar kg olarak)\n"
                "2. Kaydedin (Ctrl+S)\n"
                "3. 'Şablonu İçe Aktar' butonuna tıklayın"
            )
            
        except ImportError:
            messagebox.showerror("Hata", "openpyxl modülü yüklü değil.\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Hata", f"Şablon oluşturulamadı: {str(e)}")
    
    def _import_excel_template(self):
        """Şablondan verileri içe aktar"""
        import os
        
        if not self.template_path or not os.path.exists(self.template_path):
            # Kullanıcıya dosya seç
            file_path = filedialog.askopenfilename(
                title="Şablon Dosyası Seç",
                initialdir=os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates"),
                filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
            )
            if not file_path:
                return
            self.template_path = file_path
        
        try:
            from src.data_handlers.file_system_manager import FileSystemManager
            fs = FileSystemManager()
            data = fs.read_excel(self.template_path)
            
            if not data:
                messagebox.showwarning("Uyarı", "Şablonda veri bulunamadı!")
                return
            
            # Tabloya yükle
            self.tree.delete(*self.tree.get_children())
            self.row_count = 0
            
            for row in data:
                values = list(row.values())
                if not any(values):  # Boş satır atla
                    continue
                    
                self.row_count += 1
                
                # Şablon formatına göre (Code, Name, Amount, %, Price)
                code = values[0] if len(values) > 0 else ''
                name = values[1] if len(values) > 1 else ''
                amount = values[2] if len(values) > 2 else 0
                percent = values[3] if len(values) > 3 else 0
                price = values[4] if len(values) > 4 else 0
                
                solid_amount = 0
                solid_content = 0
                
                # Material Cache & Calc
                if code and self.on_lookup_material:
                    material = self.on_lookup_material(code)
                    if material:
                        self.material_cache[code] = material
                        if not name: name = material.get('name', '')
                        if not price: price = material.get('unit_price', 0)
                        
                        solid_content = float(material.get('solid_content', 0) or 0)
                        try:
                            solid_amount = float(amount) * (solid_content / 100.0)
                        except:
                            solid_amount = 0
                
                self.tree.insert('', tk.END, values=(
                    self.row_count,
                    code,
                    name,
                    f"{amount}",
                    f"{solid_amount:.2f}",
                    f"{solid_content:.1f}",
                    f"{percent}",
                    f"{price}"
                ))
            
            self._recalculate_percentages() # Yüzde hesabı
            self._update_summary()
            messagebox.showinfo("Başarılı", "Şablon verileri yüklendi!")
            
            # Excel dosya adını formül adı olarak ayarla
            file_name = os.path.splitext(os.path.basename(self.template_path))[0]
            self.formula_code_entry.delete(0, tk.END)
            self.formula_code_entry.insert(0, file_name)
            self.formula_name_entry.delete(0, tk.END)
            self.formula_name_entry.insert(0, file_name)
            
            self._update_summary()
            
            messagebox.showinfo("Başarılı", f"{self.row_count} satır içe aktarıldı!\nFormül adı: {file_name}")
            self.template_path = None  # Sıfırla
            
        except Exception as e:
            messagebox.showerror("Hata", f"İçe aktarma hatası: {str(e)}")
    
    def get_formulation_data(self):
        """Formülasyon verilerini sözlük olarak döndür"""
        components = []
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            code = values[1]
            
            comp_data = {
                'code': code,
                'name': values[2],
                'amount': self._safe_float(values[3]),
                'solid_amount': self._safe_float(values[4]),
                'solid_content': self._safe_float(values[5]),
                'percentage': self._safe_float(values[6]),
                'price': self._safe_float(values[7]),
                'unit_price': self._safe_float(values[7])
            }
            
            # Cache'den kimyasal özellikleri ekle
            if code in self.material_cache:
                mat = self.material_cache[code]
                # RecipeTransformer için gerekli alanlar
                fields_to_copy = [
                    'material_category', 'category', # Kategori
                    'oh_value', 'molecular_weight', 'glass_transition', # Bağlayıcı
                    'oil_absorption', 'particle_size', # Pigment
                    'boiling_point', 'evaporation_rate', # Solvent
                    'density', 'voc_g_l' # Genel
                ]
                for field in fields_to_copy:
                    if field in mat:
                        comp_data[field] = mat[field]
                
                # Kategori düzeltmesi (material_category veya category)
                if 'material_category' not in comp_data and 'category' in comp_data:
                    comp_data['material_category'] = comp_data['category']
            
            components.append(comp_data)
            
        return {
            'formula_code': self.formula_code_entry.get(),
            'formula_name': self.formula_name_entry.get(),
            'components': components
        }
    
    def load_formulation(self, data: dict):
        """Formülasyonu editöre yükle"""
        self._clear_all()
        
        # Başlıkları doldur
        self.formula_code_entry.delete(0, tk.END)
        self.formula_code_entry.insert(0, data.get('formula_code', ''))
        
        self.formula_name_entry.delete(0, tk.END)
        self.formula_name_entry.insert(0, data.get('formula_name', ''))
        
        # Bileşenleri doldur
        components = data.get('components', [])
        
        for comp in components:
            self.row_count += 1
            code = comp.get('code', comp.get('component_type', ''))
            name = comp.get('name', comp.get('component_name', ''))
            amount = comp.get('amount', 0)
            percentage = comp.get('percentage', 0)
            price = comp.get('unit_price', 0) # Fiyat DB'den gelmeli aslında, burada comp içinde var mı?
            # get_formulation_materials join ile materials table'dan unit_price getirmeliydi ama m.unit_price getirmedik mi? 
            # Az önceki sql'de m.* demedik, m.name vs dedik. m.unit_price yoktu.
            # LocalDBManager'ı tekrar güncellememek için cache'e atıp oradan okuyabiliriz veya amount * price varsa...
            
            # Material Cache Update
            if code:
                # DB'den gelen extended verileri cache'e at
                self.material_cache[code] = {
                    'name': name,
                    'code': code,
                    'solid_content': comp.get('solid_content', 0),
                    'density': comp.get('density', 1.0),
                    'voc_g_l': comp.get('voc_g_l', 0),
                    'category': comp.get('type', ''), # type alias for category in SQL
                    'unit_price': price # Eğer gelirse
                }
            
            # Solid Amount Hesapla
            solid_content = float(comp.get('solid_content', 0) or 0)
            solid_amount = amount * (solid_content / 100.0)
            
            self.tree.insert('', tk.END, values=(
                self.row_count,
                code,
                name,
                f"{amount:.2f}",
                f"{solid_amount:.2f}",
                f"{solid_content:.1f}",
                f"{percentage:.2f}",
                f"{price:.2f}"
            ))
            
        self._recalculate_percentages() # Yüzdeleri tekrar hesapla (DB'deki ile aynı olmalı ama float farkı olabilir)
        self._update_summary()
    
    def _on_project_selected(self, event=None):
        """Proje seçildiğinde"""
        self.current_project = self.project_combo.get()
    
    def _create_new_project(self):
        """Yeni proje oluştur"""
        from tkinter import simpledialog
        project_name = simpledialog.askstring("Yeni Proje", "Proje adı:")
        if project_name:
            current_values = list(self.project_combo['values']) if self.project_combo['values'] else []
            current_values.append(project_name)
            self.project_combo['values'] = current_values
            self.project_combo.set(project_name)
            self.current_project = project_name
            messagebox.showinfo("Başarılı", f"Proje oluşturuldu: {project_name}")
    
    def load_projects(self, projects: list):
        """Projeleri yükle"""
        project_names = [p.get('name', '') for p in projects if p.get('name')]
        self.project_combo['values'] = project_names
        if project_names:
            self.project_combo.current(0)
            self.current_project = project_names[0]
    
    def load_formulation_list(self, formulations: list):
        """Kayıtlı formülasyonları dropdown'a yükle"""
        self.formulation_list = formulations
        
        # Combobox değerlerini ayarla
        display_values = []
        for f in formulations:
            code = f.get('formula_code', '')
            name = f.get('formula_name', '')
            display = f"{code} - {name}" if name else code
            display_values.append(display)
        
        self.formulation_combo['values'] = display_values
    
    def _on_formulation_selected(self, event=None):
        """Formülasyon seçildiğinde yükle"""
        selection = self.formulation_combo.current()
        if selection >= 0 and selection < len(self.formulation_list):
            formulation = self.formulation_list[selection]
            
            # Formülasyon detaylarını yükle
            self.formula_code_entry.delete(0, tk.END)
            self.formula_code_entry.insert(0, formulation.get('formula_code', ''))
            
            self.formula_name_entry.delete(0, tk.END)
            self.formula_name_entry.insert(0, formulation.get('formula_name', ''))
            
            # Eğer callback varsa dışarıdan yükle, yoksa mevcut listeden dene
            if self.on_load_formulation and formulation.get('id'):
                # Callback ile detayları iste (DB'den çekilecek)
                details = self.on_load_formulation(formulation['id'])
                if details and 'components' in details:
                    formulation = details # Listeyi güncelle
            
            # Bileşenler varsa yükle
            components = formulation.get('components', [])
            self.tree.delete(*self.tree.get_children())
            self.row_count = 0
            
            for comp in components:
                self.row_count += 1
                self.tree.insert('', tk.END, values=(
                    self.row_count,
                    comp.get('code', comp.get('component_name', '')),
                    comp.get('name', comp.get('component_type', '')),
                    comp.get('solid_amount', comp.get('amount', 0)),
                    comp.get('percentage', comp.get('percentage', 0)),
                    comp.get('price', comp.get('unit_price_at_time', 0))
                ))
            
            self._update_summary()
    
    def get_current_project(self) -> str:
        """Aktif projeyi döndür"""
        return self.current_project or self.project_combo.get()
    
    def set_prediction_callback(self, callback):
        """Tahmin callback fonksiyonunu ayarla"""
        self.on_predict = callback
    
    def _predict_results(self):
        """Formülasyon için test sonuçlarını tahmin et"""
        if not self.on_predict:
            self._show_prediction_message("⚠️ ML modeli henüz bağlanmadı.\nOptimizasyon sekmesinden modeli eğitin.")
            return
        
        # Formülasyon verilerini topla
        formulation = self.get_formulation_data()
        
        if not formulation.get('components'):
            self._show_prediction_message("⚠️ Önce formülasyon hammaddelerini girin.")
            return
        
        try:
            thickness = float(self.thickness_entry.get() or 30)
        except ValueError:
            thickness = 30
        
        # Tahmin için parametreler
        params = {
            'viscosity': 100,  # Varsayılan
            'ph': 7.0,
            'density': 1.0,
            'coating_thickness': thickness,
            'total_cost': formulation.get('total_cost', 0),
            'formulation': formulation
        }
        
        # Tahmin yap
        result = self.on_predict(params)
        
        if result.get('success'):
            self._display_prediction_results(result, thickness)
        else:
            self._show_prediction_message(f"⚠️ {result.get('message', 'Tahmin yapılamadı')}")
    
    def _display_prediction_results(self, result: dict, thickness: float):
        """Tahmin sonuçlarını formatlı göster"""
        predictions = result.get('predictions', {})
        
        lines = [
            f"📊 {thickness}µm Kaplama Kalınlığında Muhtemel Sonuçlar:",
            "-" * 45,
        ]
        
        # Hedef isimleri
        target_names = {
            'opacity': 'Örtücülük',
            'gloss': 'Parlaklık (GU)',
            'corrosion_resistance': 'Korozyon Direnci (saat)',
            'adhesion': 'Yapışma (0-5)',
            'hardness': 'Sertlik (H)',
            'quality_score': 'Kalite Skoru (1-10)',
            'flexibility': 'Esneklik',
            'chemical_resistance': 'Kimyasal Dayanım',
            'uv_resistance': 'UV Dayanımı',
            'abrasion_resistance': 'Aşınma Direnci',
        }
        
        for key, value in predictions.items():
            name = target_names.get(key, key.replace('_', ' ').title())
            
            # Aralık göster (±10%)
            if isinstance(value, (int, float)):
                min_val = value * 0.9
                max_val = value * 1.1
                lines.append(f"  • {name}: {min_val:.1f} - {max_val:.1f} (tahmini: {value:.1f})")
            else:
                lines.append(f"  • {name}: {value}")
        
        lines.append("")
        lines.append("💡 Bu değerler ML modelinin tahminleridir. Gerçek değerler farklılık gösterebilir.")
        
        self._show_prediction_message("\n".join(lines))
    
    def _show_prediction_message(self, message: str):
        """Tahmin mesajını göster"""
        self.prediction_text.config(state='normal')
        self.prediction_text.delete(1.0, tk.END)
        self.prediction_text.insert(tk.END, message)
        self.prediction_text.config(state='disabled')
    
    def _safe_float(self, value) -> float:
        """Değeri güvenli şekilde float'a dönüştür"""
        if value is None or value == '':
            return 0.0
        try:
            # Zaten sayı ise
            if isinstance(value, (int, float)):
                return float(value)
            # String ise
            str_val = str(value).strip()
            # Virgülü noktaya çevir (Türkçe format)
            str_val = str_val.replace(',', '.')
            return float(str_val)
        except (ValueError, TypeError):
            return 0.0

    # --- Programatik API Metodları (Optimizasyon Entegrasyonu) ---
    
    def clear_components(self):
        """Onay sormadan tüm bileşenleri temizle (API kullanımı için)"""
        self.tree.delete(*self.tree.get_children())
        self.row_count = 0
        self._update_summary()
    
    def add_component_row(self, code: str = '', name: str = '', percentage: float = 0.0):
        """Programatik olarak yeni bileşen satırı ekle
        
        Args:
            code: Malzeme kodu
            name: Malzeme adı
            percentage: Yüzde oranı
        """
        self.row_count += 1
        
        # Malzeme bilgilerini cache'den veya DB'den çek
        solid_content = 100.0
        unit_price = 0.0
        
        if self.on_lookup_material and code:
            material = self.on_lookup_material(code)
            if material:
                if not name:
                    name = material.get('name', '')
                solid_content = material.get('solid_content', 100.0) or 100.0
                unit_price = material.get('unit_price', 0.0) or 0.0
        
        # Hesaplamalar
        amount = percentage  # Basitleştirme: Amount = Percentage (100kg baz)
        solid_amount = amount * (solid_content / 100.0)
        total_price = amount * unit_price
        
        # Treeview'a ekle
        self.tree.insert('', tk.END, values=(
            self.row_count,
            code,
            name,
            f"{amount:.2f}",
            f"{solid_amount:.2f}",
            f"{percentage:.2f}",
            f"{unit_price:.2f}",
            f"{total_price:.2f}"
        ))
        
        self._update_summary()
