"""
Paint Formulation AI - Test Sonuçları Paneli
=============================================
Formülasyonlar için test sonuçları giriş ve takip paneli
"""

import tkinter as tk
from tkinter import ttk, messagebox
from typing import Callable, Dict, List, Optional
import threading


class TestResultsPanel(ttk.LabelFrame):
    """
    Test Sonuçları Giriş Paneli
    
    Formülasyonlar için kaplama test sonuçlarını kaydetme
    """
    
    def __init__(self, parent, on_save: Callable = None, on_load_formulations: Callable = None, 
                 on_load_trial: Callable = None, on_custom_method_changed: Callable = None):
        super().__init__(parent, text="🧪 Test Sonuçları", padding=10)
        
        self.on_save = on_save
        self.on_load_formulations = on_load_formulations
        self.on_load_trial = on_load_trial  # Callback to load existing trial data
        self.on_custom_method_changed = on_custom_method_changed  # Callback when custom methods change
        
        # Formülasyon seçici
        select_frame = ttk.LabelFrame(self, text="Formülasyon Seç", padding=5)
        select_frame.pack(fill=tk.X, pady=(0, 10))
        
        row1 = ttk.Frame(select_frame)
        row1.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1, text="Proje:").pack(side=tk.LEFT)
        self.project_combo = ttk.Combobox(row1, width=25, state='readonly')
        self.project_combo.pack(side=tk.LEFT, padx=5)
        self.project_combo.bind('<<ComboboxSelected>>', self._on_project_selected)
        
        ttk.Label(row1, text="Formülasyon:").pack(side=tk.LEFT, padx=(20, 0))
        self.formulation_combo = ttk.Combobox(row1, width=25, state='readonly')
        self.formulation_combo.pack(side=tk.LEFT, padx=5)
        self.formulation_combo.bind('<<ComboboxSelected>>', self._on_formulation_selected)
        
        # Test tarihi
        row2 = ttk.Frame(select_frame)
        row2.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2, text="Test Tarihi:").pack(side=tk.LEFT)
        self.date_entry = ttk.Entry(row2, width=15)
        self.date_entry.pack(side=tk.LEFT, padx=5)
        
        from datetime import datetime
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        # Kaplama bilgileri
        coating_frame = ttk.LabelFrame(self, text="Kaplama Bilgileri", padding=5)
        coating_frame.pack(fill=tk.X, pady=5)
        
        self.coating_entries = {}
        coating_fields = [
            ("Kaplama Kalınlığı (µm):", "coating_thickness", "50"),
            ("Kuruma Süresi (dk):", "drying_time", "30"),
            ("Uygulama Metodu:", "application_method", "Fırça"),
            ("Altlık Tipi:", "substrate_type", "Çelik"),
        ]
        
        for i, (label, key, default) in enumerate(coating_fields):
            row = i // 2
            col = (i % 2) * 2
            
            ttk.Label(coating_frame, text=label).grid(row=row, column=col, sticky=tk.W, pady=2)
            entry = ttk.Entry(coating_frame, width=15)
            entry.insert(0, default)
            entry.grid(row=row, column=col+1, sticky=tk.W, padx=5, pady=2)
            self.coating_entries[key] = entry
        
        # Test sonuçları
        results_frame = ttk.LabelFrame(self, text="Test Sonuçları", padding=5)
        results_frame.pack(fill=tk.X, pady=5)
        
        self.test_entries = {}
        test_fields = [
            ("Korozyon Direnci (saat):", "corrosion_resistance", ""),
            ("Yapışma (0-5):", "adhesion", ""),
            ("Sertlik (H):", "hardness", ""),
            ("Esneklik (mm):", "flexibility", ""),
            ("Çizilme Direnci:", "scratch_resistance", ""),
            ("Aşınma Direnci:", "abrasion_resistance", ""),
            ("Kimyasal Dayanım:", "chemical_resistance", ""),
            ("UV Dayanımı:", "uv_resistance", ""),
            ("Örtücülük (%):", "opacity", ""),
            ("Parlaklık (GU):", "gloss", ""),
            ("Kalite Skoru (1-10):", "quality_score", ""),
            ("Toplam Maliyet:", "total_cost", ""),
        ]
        
        for i, (label, key, default) in enumerate(test_fields):
            row = i // 3
            col = (i % 3) * 2
            
            ttk.Label(results_frame, text=label).grid(row=row, column=col, sticky=tk.W, pady=2)
            entry = ttk.Entry(results_frame, width=12)
            if default:
                entry.insert(0, default)
            entry.grid(row=row, column=col+1, sticky=tk.W, padx=5, pady=2)
            self.test_entries[key] = entry
        
        # === ÖZEL TEST METODLARI (Test Sonuçları'nın hemen altında) ===
        custom_frame = ttk.LabelFrame(self, text="➕ Özel Test Metodları", padding=5)
        custom_frame.pack(fill=tk.X, pady=5)
        
        # Özel metodlar listesi
        self.custom_methods = {}  # key -> entry
        self.custom_methods_frame = ttk.Frame(custom_frame)
        self.custom_methods_frame.pack(fill=tk.X)
        
        # Yeni metod ekleme satırı
        add_row = ttk.Frame(custom_frame)
        add_row.pack(fill=tk.X, pady=5)
        
        ttk.Label(add_row, text="Metod Adı:").pack(side=tk.LEFT)
        self.new_method_name = ttk.Entry(add_row, width=20)
        self.new_method_name.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(add_row, text="Birim:").pack(side=tk.LEFT)
        self.new_method_unit = ttk.Entry(add_row, width=10)
        self.new_method_unit.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(add_row, text="➕ Metod Ekle", command=self._add_custom_method).pack(side=tk.LEFT, padx=5)
        
        # Önceden kaydedilmiş özel metodları yükle
        self._load_saved_custom_methods()
        
        # Notlar (Özel metodların altında)
        notes_frame = ttk.LabelFrame(self, text="Notlar / Gözlemler", padding=5)
        notes_frame.pack(fill=tk.X, pady=5)
        
        self.notes_text = tk.Text(notes_frame, height=3, wrap=tk.WORD)
        self.notes_text.pack(fill=tk.X)
        
        # === FOTOĞRAF YÜKLEME ===
        photo_frame = ttk.LabelFrame(self, text="📷 Test Fotoğrafları", padding=5)
        photo_frame.pack(fill=tk.X, pady=5)
        
        photo_btn_row = ttk.Frame(photo_frame)
        photo_btn_row.pack(fill=tk.X, pady=2)
        
        ttk.Button(photo_btn_row, text="📁 Fotoğraf Ekle", command=self._add_photo).pack(side=tk.LEFT, padx=2)
        ttk.Button(photo_btn_row, text="🗑️ Seçili Sil", command=self._remove_photo).pack(side=tk.LEFT, padx=2)
        
        self.photo_count_label = ttk.Label(photo_btn_row, text="0 fotoğraf eklendi")
        self.photo_count_label.pack(side=tk.LEFT, padx=10)
        
        # Thumbnail container
        self.photo_container = ttk.Frame(photo_frame)
        self.photo_container.pack(fill=tk.X, pady=5)
        
        self.photo_paths = []  # Store photo file paths
        self.photo_labels = []  # Store thumbnail labels
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(btn_frame, text="💾 Kaydet", command=self._save).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🧹 Alanları Temizle", command=self._clear).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📊 Geçmiş Sonuçlar", command=self._show_history).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🧠 ML Eğit", command=self._trigger_ml_training).pack(side=tk.LEFT, padx=10)
        
        ttk.Separator(btn_frame, orient='vertical').pack(side=tk.LEFT, fill=tk.Y, padx=10)
        
        ttk.Button(btn_frame, text="📝 Excel Şablonu Aç", command=self._open_excel_template).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📥 Şablonu İçe Aktar", command=self._import_excel_template).pack(side=tk.LEFT, padx=2)
        
        self.template_path = None  # Açılan şablon yolu
        
        # Geçmiş sonuçlar treeview
        history_frame = ttk.LabelFrame(self, text="Geçmiş Test Kayıtları", padding=5)
        history_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ('date', 'formulation', 'thickness', 'corrosion', 'adhesion', 'quality')
        self.history_tree = ttk.Treeview(history_frame, columns=columns, show='headings', height=6)
        
        self.history_tree.heading('date', text='Tarih')
        self.history_tree.heading('formulation', text='Formülasyon')
        self.history_tree.heading('thickness', text='Kalınlık')
        self.history_tree.heading('corrosion', text='Korozyon')
        self.history_tree.heading('adhesion', text='Yapışma')
        self.history_tree.heading('quality', text='Kalite')
        
        for col in columns:
            self.history_tree.column(col, width=80)
        
        scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)
        
        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def _on_project_selected(self, event=None):
        """Proje seçildiğinde formülasyonları yükle"""
        if self.on_load_formulations:
            project = self.project_combo.get()
            formulations = self.on_load_formulations(project)
            self.formulation_combo['values'] = formulations or []
    
    def load_projects(self, projects: list):
        """Projeleri yükle"""
        project_names = [p.get('name', '') for p in projects if p.get('name')]
        self.project_combo['values'] = project_names
    
    def load_formulations(self, formulations: list):
        """Formülasyonları yükle"""
        formula_names = [f.get('formula_code', '') or f.get('name', '') for f in formulations]
        self.formulation_combo['values'] = formula_names
    
    def _on_formulation_selected(self, event=None):
        """Formülasyon seçildiğinde mevcut test verilerini yükle"""
        formulation = self.formulation_combo.get()
        if not formulation:
            return
        
        if self.on_load_trial:
            trial_data = self.on_load_trial(formulation)
            if trial_data:
                self._fill_form_with_trial(trial_data)
                messagebox.showinfo("Bilgi", f"'{formulation}' için önceki test verileri yüklendi.")
    
    def _fill_form_with_trial(self, trial_data: dict):
        """Formu mevcut trial verileriyle doldur"""
        # Kaplama bilgilerini doldur
        coating_fields = ['coating_thickness', 'drying_time', 'application_method', 'substrate_type']
        for key in coating_fields:
            if key in self.coating_entries and key in trial_data:
                entry = self.coating_entries[key]
                entry.delete(0, tk.END)
                value = trial_data.get(key, '')
                if value is not None:
                    entry.insert(0, str(value))
        
        # Test sonuçlarını doldur
        test_fields = ['corrosion_resistance', 'adhesion', 'hardness', 'flexibility', 
                      'scratch_resistance', 'abrasion_resistance', 'chemical_resistance',
                      'uv_resistance', 'opacity', 'gloss', 'quality_score', 'total_cost']
        for key in test_fields:
            if key in self.test_entries and key in trial_data:
                entry = self.test_entries[key]
                entry.delete(0, tk.END)
                value = trial_data.get(key, '')
                if value is not None and value != '':
                    entry.insert(0, str(value))
        
        # Notları doldur
        if 'notes' in trial_data and trial_data['notes']:
            self.notes_text.delete(1.0, tk.END)
            self.notes_text.insert(tk.END, trial_data['notes'])
    
    def _save(self):
        """Test sonuçlarını kaydet"""
        data = {
            'project': self.project_combo.get(),
            'formulation': self.formulation_combo.get(),
            'date': self.date_entry.get(),
            'coating': {},
            'results': {},
            'custom_results': {},
            'notes': self.notes_text.get(1.0, tk.END).strip()
        }
        
        # Kaplama bilgileri
        for key, entry in self.coating_entries.items():
            data['coating'][key] = entry.get()
        
        # Test sonuçları (standart)
        for key, entry in self.test_entries.items():
            value = entry.get().strip()
            if value:
                try:
                    data['results'][key] = float(value)
                except ValueError:
                    data['results'][key] = value
        
        # Özel test metodları
        for key, method_data in self.custom_methods.items():
            value = method_data['entry'].get().strip()
            if value:
                try:
                    data['custom_results'][key] = float(value)
                    # Aynı zamanda results'a da ekle (ML için)
                    data['results'][key] = float(value)
                except ValueError:
                    data['custom_results'][key] = value
                    data['results'][key] = value
        
        if not data['formulation']:
            messagebox.showwarning("Uyarı", "Formülasyon seçmelisiniz!")
            return
        
        if self.on_save:
            self.on_save(data)
            
            # Geçmişe ekle
            self.history_tree.insert('', 0, values=(
                data['date'],
                data['formulation'],
                data['coating'].get('coating_thickness', ''),
                data['results'].get('corrosion_resistance', ''),
                data['results'].get('adhesion', ''),
                data['results'].get('quality_score', '')
            ))
            
            # Kayıt başarılı uyarısı - devam etme opsiyonu ile
            messagebox.showinfo("Başarılı", "✅ Test sonuçları kaydedildi!\n\nYeni kayıt girmek için değerleri değiştirebilir veya 'Alanları Temizle' butonunu kullanabilirsiniz.")
    
    def _clear(self):
        """Formu temizle"""
        for entry in self.coating_entries.values():
            entry.delete(0, tk.END)
        
        for entry in self.test_entries.values():
            entry.delete(0, tk.END)
        
        self.notes_text.delete(1.0, tk.END)
        
        # Varsayılanları geri yükle
        self.coating_entries['coating_thickness'].insert(0, "50")
        self.coating_entries['drying_time'].insert(0, "30")
        self.coating_entries['application_method'].insert(0, "Fırça")
        self.coating_entries['substrate_type'].insert(0, "Çelik")
    
    def _show_history(self):
        """Geçmiş sonuçları göster"""
        # Treeview'e scroll yap
        self.history_tree.focus_set()
        messagebox.showinfo("Bilgi", "Geçmiş test sonuçları aşağıdaki tabloda görüntüleniyor.")
    
    def _add_photo(self):
        """Fotoğraf ekle"""
        from tkinter import filedialog
        
        file_paths = filedialog.askopenfilenames(
            title="Fotoğraf Seç",
            filetypes=[
                ("Resim Dosyaları", "*.jpg *.jpeg *.png *.bmp *.gif"),
                ("Tüm Dosyalar", "*.*")
            ]
        )
        
        if not file_paths:
            return
        
        for file_path in file_paths:
            if file_path not in self.photo_paths:
                self.photo_paths.append(file_path)
                self._add_thumbnail(file_path)
        
        self._update_photo_count()
    
    def _add_thumbnail(self, path):
        """Thumbnail oluştur ve ekle"""
        import os
        
        try:
            from PIL import Image, ImageTk
            
            # Küçük resim oluştur
            img = Image.open(path)
            img.thumbnail((80, 80))
            photo = ImageTk.PhotoImage(img)
            
            # Frame oluştur
            frame = ttk.Frame(self.photo_container)
            frame.pack(side=tk.LEFT, padx=2, pady=2)
            
            # Resmi göster
            label = ttk.Label(frame, image=photo)
            label.image = photo  # Referans tut
            label.pack()
            
            # Dosya adı
            name = os.path.basename(path)[:12] + "..." if len(os.path.basename(path)) > 12 else os.path.basename(path)
            ttk.Label(frame, text=name, font=('Helvetica', 7)).pack()
            
            # Click event - seçim için
            frame.bind('<Button-1>', lambda e, p=path: self._select_photo(p))
            label.bind('<Button-1>', lambda e, p=path: self._select_photo(p))
            
            self.photo_labels.append((frame, path))
            
        except ImportError:
            # PIL yoksa sadece dosya adı göster
            frame = ttk.Frame(self.photo_container)
            frame.pack(side=tk.LEFT, padx=2, pady=2)
            
            import os
            name = os.path.basename(path)[:15]
            ttk.Label(frame, text=f"📷 {name}").pack()
            
            frame.bind('<Button-1>', lambda e, p=path: self._select_photo(p))
            self.photo_labels.append((frame, path))
        except Exception as e:
            messagebox.showerror("Hata", f"Thumbnail oluşturulamadı: {str(e)}")
    
    def _select_photo(self, path):
        """Fotoğraf seç (silme için)"""
        self.selected_photo = path
        # Seçili olanı vurgula
        for frame, p in self.photo_labels:
            if p == path:
                frame.configure(style='Selected.TFrame')
            else:
                frame.configure(style='TFrame')
    
    def _remove_photo(self):
        """Seçili fotoğrafı kaldır"""
        if not hasattr(self, 'selected_photo') or not self.selected_photo:
            messagebox.showwarning("Uyarı", "Silmek için bir fotoğraf seçin!")
            return
        
        # Listeden kaldır
        if self.selected_photo in self.photo_paths:
            self.photo_paths.remove(self.selected_photo)
        
        # UI'dan kaldır
        for frame, path in self.photo_labels[:]:
            if path == self.selected_photo:
                frame.destroy()
                self.photo_labels.remove((frame, path))
        
        self.selected_photo = None
        self._update_photo_count()
    
    def _update_photo_count(self):
        """Fotoğraf sayısını güncelle"""
        count = len(self.photo_paths)
        self.photo_count_label.config(text=f"{count} fotoğraf eklendi")
    
    def _clear_photos(self):
        """Tüm fotoğrafları temizle"""
        for frame, _ in self.photo_labels:
            frame.destroy()
        self.photo_labels.clear()
        self.photo_paths.clear()
        self._update_photo_count()
    
    def load_history(self, trials: list):
        """Geçmiş test sonuçlarını tabloya yükle"""
        # Mevcut kayıtları temizle
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)
        
        # Yeni kayıtları ekle
        for t in trials:
            trial_date = t.get('trial_date', '')
            if trial_date and len(trial_date) > 10:
                trial_date = trial_date[:10]
            
            formula = t.get('formula_code') or t.get('formula_name') or '-'
            
            self.history_tree.insert('', tk.END, values=(
                trial_date,
                formula,
                t.get('coating_thickness', '-'),
                t.get('corrosion_resistance', '-'),
                t.get('adhesion', '-'),
                t.get('quality_score', '-')
            ))
    
    def _open_excel_template(self):
        """Test sonuçları için Excel şablonu oluştur ve aç"""
        import os
        from datetime import datetime
        
        # Şablon klasörü
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
        os.makedirs(template_dir, exist_ok=True)
        
        # Benzersiz dosya adı
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        template_name = f"test_results_{timestamp}.xlsx"
        template_path = os.path.join(template_dir, template_name)
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Sonuçları"
            
            # Başlık stili
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Sütun başlıkları
            headers = ["Tarih", "Formülasyon", "Kalınlık (µm)", "Korozyon", "Yapışma", "Sertlik", "Esneklik", "Kalite (1-10)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 12
            
            # Örnek satırlar (boş) - bugünün tarihi ile
            from datetime import date
            today = date.today().isoformat()
            for row in range(2, 22):  # 20 boş satır
                ws.cell(row=row, column=1, value=today).border = thin_border
                for col in range(2, 9):
                    cell = ws.cell(row=row, column=col, value="")
                    cell.border = thin_border
            
            wb.save(template_path)
            
            # Excel'i aç
            os.startfile(template_path)
            
            self.template_path = template_path
            messagebox.showinfo(
                "Excel Şablonu", 
                f"Excel şablonu açıldı:\n{template_name}\n\n"
                "1. Test sonuçlarını girin\n"
                "2. Kaydedin (Ctrl+S)\n"
                "3. 'Şablonu İçe Aktar' butonuna tıklayın"
            )
            
        except ImportError:
            messagebox.showerror("Hata", "openpyxl modülü yüklü değil.\npip install openpyxl")
        except Exception as e:
            messagebox.showerror("Hata", f"Şablon oluşturulamadı: {str(e)}")
    
    def _import_excel_template(self):
        """Şablondan test sonuçlarını içe aktar"""
        import os
        from tkinter import filedialog
        
        if not self.template_path or not os.path.exists(self.template_path):
            # Kullanıcıya dosya seç
            file_path = filedialog.askopenfilename(
                title="Şablon Dosyası Seç",
                initialdir=os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates"),
                filetypes=[("Excel Dosyaları", "*.xlsx *.xls")]
            )
            if not file_path:
                return
            self.template_path = file_path
        
        try:
            from src.data_handlers.file_system_manager import FileSystemManager
            fs = FileSystemManager()
            data = fs.read_excel(self.template_path)
            
            if not data:
                messagebox.showwarning("Uyarı", "Şablonda veri bulunamadı!")
                return
            
            # Geçmiş tablosuna ekle
            imported_count = 0
            for row in data:
                values = list(row.values())
                if not any(values):  # Boş satır atla
                    continue
                
                trial_date = values[0] if len(values) > 0 else ''
                formula = values[1] if len(values) > 1 else ''
                thickness = values[2] if len(values) > 2 else '-'
                corrosion = values[3] if len(values) > 3 else '-'
                adhesion = values[4] if len(values) > 4 else '-'
                quality = values[7] if len(values) > 7 else '-'
                
                self.history_tree.insert('', tk.END, values=(
                    trial_date, formula, thickness, corrosion, adhesion, quality
                ))
                imported_count += 1
            
            # Excel dosya adını log olarak kullan
            file_name = os.path.splitext(os.path.basename(self.template_path))[0]
            
            messagebox.showinfo("Başarılı", f"{imported_count} test sonucu içe aktarıldı!\nGrup: {file_name}")
            self.template_path = None  # Sıfırla
            
        except Exception as e:
            messagebox.showerror("Hata", f"İçe aktarma hatası: {str(e)}")
    
    def get_test_data(self) -> dict:
        """Tüm test verilerini al"""
        data = {
            'coating': {},
            'results': {},
            'custom_methods': {}
        }
        
        for key, entry in self.coating_entries.items():
            data['coating'][key] = entry.get()
        
        for key, entry in self.test_entries.items():
            value = entry.get().strip()
            if value:
                try:
                    data['results'][key] = float(value)
                except ValueError:
                    data['results'][key] = value
        
        # Özel metodlar
        for key, method_data in self.custom_methods.items():
            value = method_data['entry'].get().strip()
            if value:
                try:
                    data['custom_methods'][key] = float(value)
                except ValueError:
                    data['custom_methods'][key] = value
        
        return data
    
    def _add_custom_method(self):
        """Yeni özel test metodu ekle"""
        name = self.new_method_name.get().strip()
        unit = self.new_method_unit.get().strip() or ""
        
        if not name:
            messagebox.showwarning("Uyarı", "Metod adı girilmelidir!")
            return
        
        # Key oluştur (küçük harf, alt çizgi)
        key = name.lower().replace(" ", "_").replace("-", "_")
        
        if key in self.custom_methods:
            messagebox.showwarning("Uyarı", "Bu metod zaten eklenmiş!")
            return
        
        # UI'a ekle
        self._add_custom_method_ui(key, name, unit)
        
        # Giriş alanlarını temizle
        self.new_method_name.delete(0, tk.END)
        self.new_method_unit.delete(0, tk.END)
        
        # Özel metodları kaydet
        self._save_custom_methods()
        
        messagebox.showinfo("Başarılı", f"'{name}' metodu eklendi. ML modeli sonraki eğitimde bu metodu kullanacak.")
    
    def _add_custom_method_ui(self, key: str, name: str, unit: str):
        """Özel metod için UI elementi ekle"""
        row = ttk.Frame(self.custom_methods_frame)
        row.pack(fill=tk.X, pady=2)
        
        label_text = f"{name} ({unit}):" if unit else f"{name}:"
        ttk.Label(row, text=label_text, width=25).pack(side=tk.LEFT)
        
        entry = ttk.Entry(row, width=12)
        entry.pack(side=tk.LEFT, padx=5)
        
        # Sil butonu
        def delete_method():
            self._delete_custom_method(key, row)
        
        ttk.Button(row, text="🗑️", width=3, command=delete_method).pack(side=tk.LEFT, padx=2)
        
        self.custom_methods[key] = {
            'name': name,
            'unit': unit,
            'entry': entry,
            'row': row
        }
    
    def _delete_custom_method(self, key: str, row):
        """Özel metodu sil"""
        if messagebox.askyesno("Onay", "Bu test metodunu silmek istiyor musunuz?"):
            row.destroy()
            del self.custom_methods[key]
            self._save_custom_methods()
    
    def _save_custom_methods(self):
        """Özel metodları kaydet (dosyaya)"""
        import json
        import os
        
        methods = {key: {'name': v['name'], 'unit': v['unit']} 
                   for key, v in self.custom_methods.items()}
        
        config_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_file = os.path.join(config_dir, 'data_storage', 'custom_test_methods.json')
        
        os.makedirs(os.path.dirname(config_file), exist_ok=True)
        
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(methods, f, ensure_ascii=False, indent=2)
        
        # Optimizasyon panelini güncelle (varsa)
        if self.on_custom_method_changed:
            self.on_custom_method_changed()
    
    def _load_saved_custom_methods(self):
        """Kaydedilmiş özel metodları yükle"""
        import json
        import os
        
        config_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_file = os.path.join(config_dir, 'data_storage', 'custom_test_methods.json')
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    methods = json.load(f)
                
                for key, data in methods.items():
                    self._add_custom_method_ui(key, data['name'], data.get('unit', ''))
            except Exception:
                pass
    
    def _trigger_ml_training(self):
        """ML eğitimini tetikle"""
        messagebox.showinfo(
            "ML Eğitimi",
            "ML modeli yeni test metodları dahil edilerek eğitilecek.\n\n"
            "Optimizasyon sekmesinden 'Modeli Eğit' butonuna tıklayın."
        )
    
    def get_all_method_keys(self) -> list:
        """Tüm test metodu anahtarlarını al (standart + özel)"""
        keys = list(self.test_entries.keys())
        keys.extend(list(self.custom_methods.keys()))
        return keys

