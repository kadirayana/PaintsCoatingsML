"""
Paint Formulation AI - Trial List Dialog
=========================================
Test sonuçları listesi popup penceresi
"""

import tkinter as tk
from tkinter import ttk
from typing import List, Dict


class TrialListDialog(tk.Toplevel):
    """
    Test sonuçları listesi popup penceresi
    
    Test sonuçlarını tablo halinde gösterir.
    """
    
    def __init__(self, parent, title: str, trials: List[Dict]):
        """
        Args:
            parent: Üst pencere
            title: Dialog başlığı
            trials: Test sonuçları listesi
        """
        super().__init__(parent)
        self.title(title)
        self.geometry("950x550")
        self.transient(parent)
        self.grab_set()
        
        self.trials = trials
        
        self._create_widgets()
        self._load_trials()
    
    def _create_widgets(self):
        """Widget'ları oluştur"""
        # Başlık
        header_frame = ttk.Frame(self, padding=10)
        header_frame.pack(fill=tk.X)
        
        ttk.Label(
            header_frame,
            text=self.title(),
            font=("Helvetica", 14, "bold")
        ).pack(side=tk.LEFT)
        
        ttk.Label(
            header_frame,
            text=f"Toplam: {len(self.trials)} test sonucu"
        ).pack(side=tk.RIGHT)
        
        # Treeview
        tree_frame = ttk.Frame(self, padding=(10, 0, 10, 10))
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ('id', 'date', 'formula', 'thickness', 'opacity', 
                   'gloss', 'hardness', 'adhesion', 'corrosion')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=18)
        
        # Sütun başlıkları
        self.tree.heading('id', text='ID')
        self.tree.heading('date', text='Tarih')
        self.tree.heading('formula', text='Formülasyon')
        self.tree.heading('thickness', text='Kalınlık (µm)')
        self.tree.heading('opacity', text='Örtücülük (%)')
        self.tree.heading('gloss', text='Parlaklık')
        self.tree.heading('hardness', text='Sertlik')
        self.tree.heading('adhesion', text='Yapışma')
        self.tree.heading('corrosion', text='Korozyon')
        
        # Sütun genişlikleri
        self.tree.column('id', width=40, anchor=tk.CENTER)
        self.tree.column('date', width=90, anchor=tk.CENTER)
        self.tree.column('formula', width=150)
        self.tree.column('thickness', width=90, anchor=tk.CENTER)
        self.tree.column('opacity', width=90, anchor=tk.CENTER)
        self.tree.column('gloss', width=80, anchor=tk.CENTER)
        self.tree.column('hardness', width=80, anchor=tk.CENTER)
        self.tree.column('adhesion', width=80, anchor=tk.CENTER)
        self.tree.column('corrosion', width=80, anchor=tk.CENTER)
        
        # Scrollbar
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        
        # Butonlar
        btn_frame = ttk.Frame(self, padding=10)
        btn_frame.pack(fill=tk.X)
        
        ttk.Button(
            btn_frame,
            text="📊 İstatistikler",
            command=self._show_statistics
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="📋 Excel'e Aktar",
            command=self._export_to_excel
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="Kapat",
            command=self.destroy
        ).pack(side=tk.RIGHT, padx=5)
    
    def _load_trials(self):
        """Test sonuçlarını tabloya yükle"""
        for t in self.trials:
            trial_date = t.get('test_date', '') or t.get('trial_date', '')
            if trial_date and len(trial_date) > 10:
                trial_date = trial_date[:10]
            
            formula = (t.get('formula_code') or 
                      t.get('formulation_code') or 
                      t.get('formula_name') or '-')
            
            self.tree.insert('', tk.END, values=(
                t.get('id', ''),
                trial_date,
                formula,
                self._format_value(t.get('film_thickness')),
                self._format_value(t.get('opacity')),
                self._format_value(t.get('gloss_60') or t.get('gloss')),
                self._format_value(t.get('hardness')),
                t.get('adhesion', '-'),
                t.get('corrosion_resistance', '-')
            ))
    
    def _format_value(self, value) -> str:
        """Değeri formatlı stringe çevir"""
        if value is None:
            return '-'
        try:
            return f"{float(value):.1f}"
        except (ValueError, TypeError):
            return str(value)
    
    def _show_statistics(self):
        """İstatistik özeti göster"""
        if not self.trials:
            return
        
        # Basit istatistikler hesapla
        values = {
            'film_thickness': [],
            'opacity': [],
            'gloss_60': [],
            'hardness': []
        }
        
        for t in self.trials:
            for key in values:
                val = t.get(key)
                if val is not None:
                    try:
                        values[key].append(float(val))
                    except:
                        pass
        
        # İstatistik penceresi
        stats_window = tk.Toplevel(self)
        stats_window.title("İstatistikler")
        stats_window.geometry("300x250")
        stats_window.transient(self)
        
        ttk.Label(
            stats_window,
            text="Test Sonucu İstatistikleri",
            font=("Helvetica", 12, "bold")
        ).pack(pady=10)
        
        stats_frame = ttk.Frame(stats_window, padding=10)
        stats_frame.pack(fill=tk.BOTH, expand=True)
        
        labels = {
            'film_thickness': 'Film Kalınlığı',
            'opacity': 'Örtücülük',
            'gloss_60': 'Parlaklık',
            'hardness': 'Sertlik'
        }
        
        for i, (key, name) in enumerate(labels.items()):
            vals = values[key]
            if vals:
                avg = sum(vals) / len(vals)
                min_val = min(vals)
                max_val = max(vals)
                text = f"{name}: Ort={avg:.1f}, Min={min_val:.1f}, Max={max_val:.1f}"
            else:
                text = f"{name}: Veri yok"
            
            ttk.Label(stats_frame, text=text).pack(anchor=tk.W, pady=2)
        
        ttk.Button(stats_window, text="Kapat", command=stats_window.destroy).pack(pady=10)
    
    def _export_to_excel(self):
        """Excel'e aktar"""
        from tkinter import filedialog, messagebox
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Test Sonuçlarını Kaydet"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Sonuçları"
            
            # Başlıklar
            headers = ['ID', 'Tarih', 'Formül', 'Kalınlık', 'Örtücülük', 
                      'Parlaklık', 'Sertlik', 'Yapışma', 'Korozyon']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Veriler
            for row, item in enumerate(self.tree.get_children(), 2):
                values = self.tree.item(item)['values']
                for col, value in enumerate(values, 1):
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(file_path)
            messagebox.showinfo("Başarılı", f"Dosya kaydedildi:\n{file_path}")
            
        except ImportError:
            messagebox.showerror("Hata", "openpyxl kütüphanesi bulunamadı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya kaydedilemedi: {e}")
