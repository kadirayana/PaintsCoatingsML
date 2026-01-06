"""
Paint Formulation AI - Excel Handler
=====================================
Excel import/export işlemleri
"""

import os
import tkinter as tk
from tkinter import filedialog, messagebox
from typing import Dict, List, Optional, Callable
import logging

logger = logging.getLogger(__name__)


class ExcelHandler:
    """
    Excel dosya işlemleri yöneticisi
    
    Formülasyonları Excel'den okuma ve Excel'e yazma işlemlerini yönetir.
    """
    
    # Excel sütun eşlemeleri (Türkçe başlıklar)
    COLUMN_MAPPINGS = {
        'code': ['Hammadde Kodu', 'Malzeme Kodu', 'Kod', 'Code', 'Material Code'],
        'name': ['Hammadde Adı', 'Malzeme Adı', 'Ad', 'Name', 'Material Name'],
        'amount': ['Miktar', 'Amount', 'Kg', 'Quantity'],
        'solid_content': ['Katı %', 'Katı Oranı', 'Solid %', 'Solid Content'],
        'unit_price': ['Birim Fiyat', 'Fiyat', 'Price', 'Unit Price'],
    }
    
    def __init__(self, parent_window=None):
        """
        Args:
            parent_window: Üst pencere (dialog parent)
        """
        self.parent = parent_window
    
    def import_from_excel(self, file_path: str = None) -> Optional[List[Dict]]:
        """
        Excel dosyasından formülasyon verilerini oku
        
        Args:
            file_path: Dosya yolu (None ise dialog açılır)
            
        Returns:
            List[Dict]: Bileşen verileri veya None
        """
        if file_path is None:
            file_path = filedialog.askopenfilename(
                parent=self.parent,
                title="Excel Dosyası Seç",
                filetypes=[
                    ("Excel dosyaları", "*.xlsx *.xls"),
                    ("Tüm dosyalar", "*.*")
                ]
            )
        
        if not file_path:
            return None
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Başlık satırını bul
            headers = {}
            header_row = 1
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    header_text = str(cell.value).strip()
                    # Eşleşen alan bul
                    for field, alternatives in self.COLUMN_MAPPINGS.items():
                        if header_text in alternatives:
                            headers[field] = col_idx
                            break
            
            if not headers:
                messagebox.showerror("Hata", "Excel başlıkları tanınamadı.")
                return None
            
            # Veri satırlarını oku
            components = []
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                # Boş satırları atla
                if all(cell.value is None for cell in row):
                    continue
                
                component = {}
                for field, col_idx in headers.items():
                    cell_value = row[col_idx - 1].value
                    component[field] = cell_value if cell_value else ''
                
                # En az kod veya isim olmalı
                if component.get('code') or component.get('name'):
                    # Hesaplamaları yap
                    self._calculate_component(component)
                    components.append(component)
            
            wb.close()
            
            logger.info(f"Excel'den {len(components)} bileşen yüklendi: {file_path}")
            return components
            
        except ImportError:
            messagebox.showerror("Hata", "openpyxl kütüphanesi bulunamadı.\npip install openpyxl")
            return None
        except Exception as e:
            logger.error(f"Excel okuma hatası: {e}")
            messagebox.showerror("Hata", f"Excel okunamadı:\n{str(e)}")
            return None
    
    def export_to_excel(self, components: List[Dict], file_path: str = None) -> bool:
        """
        Formülasyon verilerini Excel'e yaz
        
        Args:
            components: Bileşen verileri
            file_path: Hedef dosya yolu (None ise dialog açılır)
            
        Returns:
            bool: Başarılı ise True
        """
        if file_path is None:
            file_path = filedialog.asksaveasfilename(
                parent=self.parent,
                title="Excel Olarak Kaydet",
                defaultextension=".xlsx",
                filetypes=[("Excel dosyası", "*.xlsx")]
            )
        
        if not file_path:
            return False
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Formülasyon"
            
            # Başlık stili
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Başlıklar
            headers = [
                'Hammadde Kodu', 'Hammadde Adı', 'Miktar (kg)', 
                'Katı %', 'Katı (kg)', '%', 'Birim Fiyat', 'Toplam Fiyat'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # Veriler
            for row_idx, comp in enumerate(components, 2):
                values = [
                    comp.get('code', ''),
                    comp.get('name', ''),
                    comp.get('amount', ''),
                    comp.get('solid_content', ''),
                    comp.get('solid_amount', ''),
                    comp.get('percentage', ''),
                    comp.get('unit_price', ''),
                    comp.get('total_price', '')
                ]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            
            wb.save(file_path)
            wb.close()
            
            logger.info(f"Excel'e {len(components)} bileşen kaydedildi: {file_path}")
            messagebox.showinfo("Başarılı", f"Formülasyon Excel'e kaydedildi:\n{file_path}")
            return True
            
        except ImportError:
            messagebox.showerror("Hata", "openpyxl kütüphanesi bulunamadı.")
            return False
        except Exception as e:
            logger.error(f"Excel yazma hatası: {e}")
            messagebox.showerror("Hata", f"Excel kaydedilemedi:\n{str(e)}")
            return False
    
    def create_template(self, file_path: str = None) -> Optional[str]:
        """
        Boş Excel şablonu oluştur
        
        Args:
            file_path: Hedef dosya yolu (None ise temp dizine)
            
        Returns:
            str: Oluşturulan dosya yolu
        """
        if file_path is None:
            import tempfile
            file_path = os.path.join(tempfile.gettempdir(), "formulation_template.xlsx")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Formülasyon Şablonu"
            
            # Başlıklar
            headers = [
                'Hammadde Kodu', 'Hammadde Adı', 'Miktar (kg)', 
                'Katı %', 'Birim Fiyat'
            ]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Örnek satırlar
            example_data = [
                ('RES001', 'Epoksi Reçine', 25, 100, 45.50),
                ('PIG001', 'Titanyum Dioksit', 15, 100, 82.00),
                ('SOL001', 'Ksilen', 10, 0, 12.30),
            ]
            
            for row_idx, data in enumerate(example_data, 2):
                for col, value in enumerate(data, 1):
                    ws.cell(row=row_idx, column=col, value=value)
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            
            wb.save(file_path)
            wb.close()
            
            logger.info(f"Excel şablonu oluşturuldu: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Şablon oluşturma hatası: {e}")
            return None
    
    def _calculate_component(self, component: Dict):
        """Bileşen hesaplamalarını yap"""
        try:
            amount = float(component.get('amount', 0) or 0)
            solid_content = float(component.get('solid_content', 100) or 100)
            unit_price = float(component.get('unit_price', 0) or 0)
            
            # Katı miktarı
            component['solid_amount'] = amount * solid_content / 100
            
            # Toplam fiyat
            component['total_price'] = amount * unit_price
            
        except (ValueError, TypeError):
            pass
