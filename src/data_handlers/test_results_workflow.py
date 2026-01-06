"""
Paint Formulation AI - Test Results Excel Workflow
==================================================
Advanced Excel import/export for Test Results with validation and atomic transactions.

Features:
- Smart template generation with Excel-side data validation
- Formulation code linkage check
- Atomic transactions with duplicate handling
- Pre-filled templates for untested formulations
"""

import os
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Callable, Any
from dataclasses import dataclass, field

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TestImportRow:
    """A single row from test results import"""
    row_number: int
    formulation_code: str
    formulation_id: Optional[int] = None
    test_date: Optional[datetime] = None
    thickness: Optional[float] = None
    gloss: Optional[float] = None
    corrosion_resistance: Optional[float] = None
    adhesion: Optional[float] = None
    hardness: Optional[float] = None
    flexibility: Optional[float] = None
    opacity: Optional[float] = None
    quality_score: Optional[float] = None
    viscosity: Optional[float] = None
    ph: Optional[float] = None
    density: Optional[float] = None
    notes: str = ""
    
    # Validation status
    is_valid: bool = True
    validation_errors: List[str] = field(default_factory=list)
    is_duplicate: bool = False
    existing_trial_id: Optional[int] = None


@dataclass
class TestImportValidationResult:
    """Result of test import validation"""
    total_rows: int = 0
    valid_rows: int = 0
    invalid_rows: int = 0
    duplicate_rows: int = 0
    orphaned_rows: int = 0
    rows: List[TestImportRow] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    
    @property
    def can_import(self) -> bool:
        return self.valid_rows > 0


@dataclass
class TestImportResult:
    """Result of test import operation"""
    success: bool
    message: str
    imported_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    rollback_performed: bool = False


# ============================================================================
# PHYSICAL LIMITS FOR SANITY CHECKS
# ============================================================================

PHYSICAL_LIMITS = {
    'opacity': (0, 100),           # Percentage
    'gloss': (0, 200),             # GU - can exceed 100 for high gloss
    'adhesion': (0, 5),            # Scale 0-5
    'quality_score': (1, 10),      # Scale 1-10
    'hardness': (0, 10),           # Pencil hardness scale approximate
    'flexibility': (0, 10),        # Scale
    'corrosion_resistance': (0, 10000),  # Hours
    'thickness': (0, 5000),        # Micrometers
    'viscosity': (1, 100000),      # cP
    'ph': (0, 14),                 # pH scale
    'density': (0.1, 5)            # g/cm³
}


# ============================================================================
# SMART TEMPLATE GENERATOR FOR TEST RESULTS
# ============================================================================

class TestResultsTemplateGenerator:
    """
    Generates Excel templates for test results with:
    - Hidden metadata sheet
    - Excel-side data validation
    - Pre-filled formulation codes
    """
    
    METADATA_SHEET_NAME = "_SYSTEM_METADATA"
    DATA_SHEET_NAME = "Test_Results"
    FORMULATIONS_SHEET_NAME = "_FORMULATION_CODES"
    
    # Column definitions: (key, header, width, required)
    COLUMNS = [
        ('formulation_code', 'Formülasyon Kodu *', 18, True),
        ('test_date', 'Test Tarihi *', 14, True),
        ('thickness', 'Kalınlık (µm)', 12, False),
        ('gloss', 'Parlaklık (GU)', 14, False),
        ('corrosion_resistance', 'Korozyon (saat)', 14, False),
        ('adhesion', 'Yapışma (0-5)', 14, False),
        ('hardness', 'Sertlik', 10, False),
        ('flexibility', 'Esneklik', 10, False),
        ('opacity', 'Örtücülük (%)', 14, False),
        ('quality_score', 'Kalite (1-10)', 14, False),
        ('viscosity', 'Viskozite (cP)', 14, False),
        ('ph', 'pH', 8, False),
        ('density', 'Yoğunluk (g/cm³)', 14, False),
        ('notes', 'Notlar', 30, False)
    ]
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def generate_template(
        self,
        project_id: Optional[int] = None,
        project_name: str = "",
        output_path: Optional[str] = None,
        prefill_untested: bool = True
    ) -> Tuple[Any, str]:
        """
        Generate a smart test results template.
        
        Args:
            project_id: Project to scope formulations
            project_name: Project name for filename
            output_path: Optional save path
            prefill_untested: Pre-fill with untested formulation codes
            
        Returns:
            Tuple of (Workbook, suggested_filename)
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl kütüphanesi gerekli")
        
        wb = Workbook()
        
        # Get formulation codes for this project
        formulation_codes = self._get_project_formulation_codes(project_id)
        untested_codes = self._get_untested_formulation_codes(project_id) if prefill_untested else []
        
        # Create sheets
        self._create_data_sheet(wb, untested_codes)
        self._create_metadata_sheet(wb, project_id, project_name)
        self._create_formulations_sheet(wb, formulation_codes)
        
        # Add data validations
        self._add_data_validations(wb, len(formulation_codes))
        
        # Generate filename
        filename = self._generate_filename(project_name)
        
        if output_path:
            if os.path.isdir(output_path):
                file_path = os.path.join(output_path, filename)
            else:
                file_path = output_path
            wb.save(file_path)
            return wb, file_path
        
        return wb, filename
    
    def _create_data_sheet(self, wb: Workbook, prefill_codes: List[str]):
        """Create the main data entry sheet"""
        ws = wb.active
        ws.title = self.DATA_SHEET_NAME
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, (key, header, width, required) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Pre-fill untested formulation codes
        if prefill_codes:
            for row_idx, code in enumerate(prefill_codes, 2):
                ws.cell(row=row_idx, column=1, value=code)
                # Set today's date as default
                ws.cell(row=row_idx, column=2, value=datetime.now().date())
        
        # Add some empty rows for data entry
        if not prefill_codes:
            for row_idx in range(2, 12):
                ws.cell(row=row_idx, column=2, value=datetime.now().date())
    
    def _create_metadata_sheet(self, wb: Workbook, project_id: Optional[int], project_name: str):
        """Create hidden metadata sheet"""
        ws = wb.create_sheet(self.METADATA_SHEET_NAME)
        
        metadata = {
            'Project_ID': project_id,
            'Project_Name': project_name,
            'Template_Type': 'TestResults',
            'Template_Version': 'v2.0',
            'Export_Timestamp': datetime.now().isoformat(),
            'Generated_By': 'PaintFormulationAI'
        }
        
        for row, (key, value) in enumerate(metadata.items(), 1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value) if value is not None else '')
        
        # Hide the sheet
        ws.sheet_state = 'hidden'
    
    def _create_formulations_sheet(self, wb: Workbook, codes: List[str]):
        """Create hidden sheet with formulation codes for dropdown"""
        ws = wb.create_sheet(self.FORMULATIONS_SHEET_NAME)
        
        for row, code in enumerate(codes, 1):
            ws.cell(row=row, column=1, value=code)
        
        # Hide the sheet
        ws.sheet_state = 'hidden'
    
    def _add_data_validations(self, wb: Workbook, formulation_count: int):
        """Add Excel-side data validations"""
        ws = wb[self.DATA_SHEET_NAME]
        
        # Adhesion: 0-5 integer
        adhesion_dv = DataValidation(
            type="whole",
            operator="between",
            formula1=0,
            formula2=5,
            allow_blank=True
        )
        adhesion_dv.error = "Yapışma değeri 0-5 arasında olmalıdır"
        adhesion_dv.errorTitle = "Geçersiz Değer"
        ws.add_data_validation(adhesion_dv)
        adhesion_dv.add('F2:F1000')  # Column F = Adhesion
        
        # Quality Score: 1-10 integer
        quality_dv = DataValidation(
            type="whole",
            operator="between",
            formula1=1,
            formula2=10,
            allow_blank=True
        )
        quality_dv.error = "Kalite skoru 1-10 arasında olmalıdır"
        quality_dv.errorTitle = "Geçersiz Değer"
        ws.add_data_validation(quality_dv)
        quality_dv.add('J2:J1000')  # Column J = Quality Score
        
        # Opacity: 0-100 percentage
        opacity_dv = DataValidation(
            type="decimal",
            operator="between",
            formula1=0,
            formula2=100,
            allow_blank=True
        )
        opacity_dv.error = "Örtücülük %0-100 arasında olmalıdır"
        opacity_dv.errorTitle = "Geçersiz Değer"
        ws.add_data_validation(opacity_dv)
        opacity_dv.add('I2:I1000')  # Column I = Opacity
        
        # Date format
        date_dv = DataValidation(
            type="date",
            allow_blank=True
        )
        date_dv.error = "Geçerli bir tarih girin"
        date_dv.errorTitle = "Geçersiz Tarih"
        ws.add_data_validation(date_dv)
        date_dv.add('B2:B1000')  # Column B = Test Date
        
        # Formulation Code dropdown (if we have codes)
        if formulation_count > 0:
            formula_dv = DataValidation(
                type="list",
                formula1=f"'{self.FORMULATIONS_SHEET_NAME}'!$A$1:$A${formulation_count}",
                allow_blank=False
            )
            formula_dv.error = "Geçerli bir formülasyon kodu seçin"
            formula_dv.errorTitle = "Geçersiz Formülasyon"
            formula_dv.prompt = "Listeden bir formülasyon kodu seçin"
            formula_dv.promptTitle = "Formülasyon Seç"
            ws.add_data_validation(formula_dv)
            formula_dv.add('A2:A1000')  # Column A = Formulation Code
    
    def _get_project_formulation_codes(self, project_id: Optional[int]) -> List[str]:
        """Get all formulation codes for a project"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                if project_id:
                    cursor.execute(
                        'SELECT formula_code FROM formulations WHERE project_id = ? ORDER BY formula_code',
                        (project_id,)
                    )
                else:
                    cursor.execute('SELECT formula_code FROM formulations ORDER BY formula_code')
                
                return [row['formula_code'] for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"Could not get formulation codes: {e}")
            return []
    
    def _get_untested_formulation_codes(self, project_id: Optional[int]) -> List[str]:
        """Get formulation codes that have no test results yet"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                if project_id:
                    cursor.execute('''
                        SELECT f.formula_code 
                        FROM formulations f
                        LEFT JOIN trials t ON f.id = t.formulation_id
                        WHERE f.project_id = ? AND t.id IS NULL
                        ORDER BY f.formula_code
                    ''', (project_id,))
                else:
                    cursor.execute('''
                        SELECT f.formula_code 
                        FROM formulations f
                        LEFT JOIN trials t ON f.id = t.formulation_id
                        WHERE t.id IS NULL
                        ORDER BY f.formula_code
                    ''')
                
                return [row['formula_code'] for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"Could not get untested formulations: {e}")
            return []
    
    def _generate_filename(self, project_name: str) -> str:
        """Generate filename"""
        date_str = datetime.now().strftime("%Y%m%d")
        safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_'))
        safe_name = safe_name.strip().replace(' ', '_') or "Tests"
        return f"TestResults_{safe_name}_{date_str}.xlsx"


# ============================================================================
# TEST RESULTS IMPORT HANDLER
# ============================================================================

class TestResultsImportHandler:
    """
    Handles test results import with:
    - Formulation linkage validation
    - Duplicate detection
    - Sanity checks on physical values
    - Atomic transactions
    """
    
    METADATA_SHEET_NAME = "_SYSTEM_METADATA"
    
    # Column mappings for flexible import
    COLUMN_MAPPINGS = {
        'formulation_code': ['formülasyon kodu', 'formula code', 'formulation code', 'kod', 'code'],
        'test_date': ['test tarihi', 'date', 'tarih', 'test date'],
        'thickness': ['kalınlık', 'thickness', 'coating thickness', 'kalınlık (µm)'],
        'gloss': ['parlaklık', 'gloss', 'parlaklık (gu)'],
        'corrosion_resistance': ['korozyon', 'corrosion', 'korozyon (saat)', 'corrosion resistance'],
        'adhesion': ['yapışma', 'adhesion', 'yapışma (0-5)'],
        'hardness': ['sertlik', 'hardness'],
        'flexibility': ['esneklik', 'flexibility'],
        'opacity': ['örtücülük', 'opacity', 'örtücülük (%)'],
        'quality_score': ['kalite', 'quality', 'quality score', 'kalite (1-10)'],
        'viscosity': ['viskozite', 'viscosity', 'viskozite (cp)'],
        'ph': ['ph', 'pH'],
        'density': ['yoğunluk', 'density', 'yoğunluk (g/cm³)'],
        'notes': ['notlar', 'notes', 'açıklama']
    }
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def validate_import(
        self,
        file_path: str,
        project_id: Optional[int] = None
    ) -> TestImportValidationResult:
        """
        Validate Excel file before import (the "Linkage Check").
        
        Returns validation result without making any database changes.
        """
        result = TestImportValidationResult()
        
        if not HAS_PANDAS:
            result.errors.append("Pandas kütüphanesi gerekli")
            return result
        
        try:
            # Read Excel
            df = pd.read_excel(file_path, sheet_name=0)
            
            if df.empty:
                result.errors.append("Excel dosyası boş")
                return result
            
            # Normalize columns
            df = self._normalize_columns(df)
            
            # Check required columns
            if 'formulation_code' not in df.columns:
                result.errors.append("'Formülasyon Kodu' sütunu bulunamadı")
                return result
            
            # Read metadata for project context
            file_project_id = self._read_metadata_project_id(file_path)
            
            # Use file project ID if current is not specified
            effective_project_id = project_id or file_project_id
            
            # Get formulation lookup
            formulation_lookup = self._build_formulation_lookup(effective_project_id)
            
            # Validate each row
            for idx, row in df.iterrows():
                import_row = self._parse_row(idx + 2, row)  # +2 for 1-indexed + header
                
                if not import_row.formulation_code:
                    import_row.is_valid = False
                    import_row.validation_errors.append("Formülasyon kodu boş")
                    result.rows.append(import_row)
                    continue
                
                # Linkage check
                formulation = formulation_lookup.get(import_row.formulation_code.upper())
                
                if formulation:
                    import_row.formulation_id = formulation['id']
                    
                    # Check for duplicate
                    existing = self._check_duplicate(
                        import_row.formulation_id,
                        import_row.test_date
                    )
                    if existing:
                        import_row.is_duplicate = True
                        import_row.existing_trial_id = existing['id']
                else:
                    import_row.is_valid = False
                    import_row.validation_errors.append(
                        f"Formülasyon '{import_row.formulation_code}' bu projede bulunamadı"
                    )
                    result.orphaned_rows += 1
                
                # Sanity checks
                sanity_errors = self._validate_physical_limits(import_row)
                if sanity_errors:
                    import_row.validation_errors.extend(sanity_errors)
                    import_row.is_valid = False
                
                result.rows.append(import_row)
            
            # Count results
            result.total_rows = len(result.rows)
            result.valid_rows = sum(1 for r in result.rows if r.is_valid)
            result.invalid_rows = sum(1 for r in result.rows if not r.is_valid)
            result.duplicate_rows = sum(1 for r in result.rows if r.is_duplicate)
            
        except Exception as e:
            result.errors.append(f"Dosya okuma hatası: {str(e)}")
            logger.error(f"Validation error: {e}")
        
        return result
    
    def import_with_transaction(
        self,
        validation_result: TestImportValidationResult,
        update_duplicates: bool = False,
        on_progress: Callable[[str], None] = None
    ) -> TestImportResult:
        """
        Import validated test results with atomic transaction.
        
        Args:
            validation_result: Result from validate_import()
            update_duplicates: If True, update existing records; if False, skip
            on_progress: Progress callback
        """
        def log(msg):
            logger.info(msg)
            if on_progress:
                on_progress(msg)
        
        if not validation_result.can_import:
            return TestImportResult(
                success=False,
                message="İçe aktarılacak geçerli veri yok"
            )
        
        imported = 0
        updated = 0
        skipped = 0
        
        log("📥 Test sonuçları içe aktarılıyor...")
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                try:
                    for row in validation_result.rows:
                        if not row.is_valid:
                            skipped += 1
                            continue
                        
                        if row.is_duplicate:
                            if update_duplicates:
                                # Update existing record
                                self._update_trial(cursor, row)
                                updated += 1
                                log(f"  ↻ Güncellendi: {row.formulation_code}")
                            else:
                                skipped += 1
                                log(f"  ⏭ Atlandı (mevcut): {row.formulation_code}")
                        else:
                            # Insert new record
                            self._insert_trial(cursor, row)
                            imported += 1
                            log(f"  ✓ Eklendi: {row.formulation_code}")
                    
                    # COMMIT
                    conn.commit()
                    log(f"\n✅ Tamamlandı: {imported} eklendi, {updated} güncellendi, {skipped} atlandı")
                    
                    return TestImportResult(
                        success=True,
                        message=f"{imported + updated} test sonucu başarıyla içe aktarıldı",
                        imported_count=imported,
                        updated_count=updated,
                        skipped_count=skipped
                    )
                    
                except Exception as e:
                    # ROLLBACK
                    conn.rollback()
                    logger.error(f"Transaction rollback: {e}")
                    
                    return TestImportResult(
                        success=False,
                        message=f"İçe aktarım hatası: {str(e)}",
                        rollback_performed=True
                    )
                    
        except Exception as e:
            return TestImportResult(
                success=False,
                message=f"Veritabanı hatası: {str(e)}"
            )
    
    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize column names"""
        column_map = {}
        
        for col in df.columns:
            col_lower = str(col).lower().strip().replace('*', '').strip()
            
            for standard, variations in self.COLUMN_MAPPINGS.items():
                if col_lower in variations:
                    column_map[col] = standard
                    break
        
        return df.rename(columns=column_map)
    
    def _read_metadata_project_id(self, file_path: str) -> Optional[int]:
        """Read project ID from metadata sheet"""
        try:
            if not HAS_OPENPYXL:
                return None
            
            wb = load_workbook(file_path, read_only=True)
            
            if self.METADATA_SHEET_NAME in wb.sheetnames:
                ws = wb[self.METADATA_SHEET_NAME]
                for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                    if row[0] == 'Project_ID' and row[1]:
                        try:
                            return int(row[1])
                        except (ValueError, TypeError):
                            pass
            
            wb.close()
        except Exception as e:
            logger.warning(f"Could not read metadata: {e}")
        
        return None
    
    def _build_formulation_lookup(self, project_id: Optional[int]) -> Dict[str, Dict]:
        """Build lookup dictionary for formulations"""
        lookup = {}
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                if project_id:
                    cursor.execute(
                        'SELECT id, formula_code, formula_name FROM formulations WHERE project_id = ?',
                        (project_id,)
                    )
                else:
                    cursor.execute('SELECT id, formula_code, formula_name FROM formulations')
                
                for row in cursor.fetchall():
                    code = row['formula_code'].upper() if row['formula_code'] else ''
                    lookup[code] = dict(row)
                    
        except Exception as e:
            logger.error(f"Error building formulation lookup: {e}")
        
        return lookup
    
    def _parse_row(self, row_num: int, row: pd.Series) -> TestImportRow:
        """Parse a dataframe row into TestImportRow"""
        import_row = TestImportRow(
            row_number=row_num,
            formulation_code=str(row.get('formulation_code', '')).strip()
        )
        
        # Parse date
        date_val = row.get('test_date')
        if pd.notna(date_val):
            if isinstance(date_val, datetime):
                import_row.test_date = date_val
            elif isinstance(date_val, str):
                try:
                    import_row.test_date = pd.to_datetime(date_val)
                except Exception:
                    pass
        
        # Parse numeric fields
        numeric_fields = [
            'thickness', 'gloss', 'corrosion_resistance', 'adhesion',
            'hardness', 'flexibility', 'opacity', 'quality_score',
            'viscosity', 'ph', 'density'
        ]
        
        for field in numeric_fields:
            value = row.get(field)
            if pd.notna(value):
                try:
                    setattr(import_row, field, float(value))
                except (ValueError, TypeError):
                    pass
        
        # Parse notes
        notes = row.get('notes')
        if pd.notna(notes):
            import_row.notes = str(notes)
        
        return import_row
    
    def _validate_physical_limits(self, row: TestImportRow) -> List[str]:
        """Validate values are within physical limits"""
        errors = []
        
        for field, (min_val, max_val) in PHYSICAL_LIMITS.items():
            value = getattr(row, field, None)
            if value is not None:
                if value < min_val or value > max_val:
                    errors.append(f"{field}: {value} değeri {min_val}-{max_val} aralığında olmalı")
        
        return errors
    
    def _check_duplicate(
        self,
        formulation_id: int,
        test_date: Optional[datetime]
    ) -> Optional[Dict]:
        """Check if a trial already exists for this formulation and date"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                if test_date:
                    cursor.execute('''
                        SELECT id FROM trials 
                        WHERE formulation_id = ? AND DATE(trial_date) = DATE(?)
                    ''', (formulation_id, test_date))
                else:
                    cursor.execute('''
                        SELECT id FROM trials 
                        WHERE formulation_id = ? AND DATE(trial_date) = DATE('now')
                    ''', (formulation_id,))
                
                row = cursor.fetchone()
                return dict(row) if row else None
                
        except Exception:
            return None
    
    def _insert_trial(self, cursor, row: TestImportRow):
        """Insert a new trial record"""
        cursor.execute('''
            INSERT INTO trials (
                formulation_id, trial_date, coating_thickness, gloss,
                corrosion_resistance, adhesion, hardness, flexibility,
                opacity, quality_score, viscosity, ph, density, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            row.formulation_id,
            row.test_date or datetime.now(),
            row.thickness,
            row.gloss,
            row.corrosion_resistance,
            row.adhesion,
            row.hardness,
            row.flexibility,
            row.opacity,
            row.quality_score,
            row.viscosity,
            row.ph,
            row.density,
            row.notes
        ))
    
    def _update_trial(self, cursor, row: TestImportRow):
        """Update an existing trial record"""
        cursor.execute('''
            UPDATE trials SET
                coating_thickness = COALESCE(?, coating_thickness),
                gloss = COALESCE(?, gloss),
                corrosion_resistance = COALESCE(?, corrosion_resistance),
                adhesion = COALESCE(?, adhesion),
                hardness = COALESCE(?, hardness),
                flexibility = COALESCE(?, flexibility),
                opacity = COALESCE(?, opacity),
                quality_score = COALESCE(?, quality_score),
                viscosity = COALESCE(?, viscosity),
                ph = COALESCE(?, ph),
                density = COALESCE(?, density),
                notes = COALESCE(?, notes)
            WHERE id = ?
        ''', (
            row.thickness,
            row.gloss,
            row.corrosion_resistance,
            row.adhesion,
            row.hardness,
            row.flexibility,
            row.opacity,
            row.quality_score,
            row.viscosity,
            row.ph,
            row.density,
            row.notes,
            row.existing_trial_id
        ))


# ============================================================================
# INTEGRATED WORKFLOW
# ============================================================================

class TestResultsExcelWorkflow:
    """
    High-level workflow controller for test results Excel operations.
    """
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.template_generator = TestResultsTemplateGenerator(db_manager)
        self.import_handler = TestResultsImportHandler(db_manager)
    
    def download_template(
        self,
        project_id: Optional[int] = None,
        project_name: str = "",
        prefill_untested: bool = True
    ) -> Tuple[Any, str]:
        """Get template workbook for download"""
        return self.template_generator.generate_template(
            project_id=project_id,
            project_name=project_name,
            prefill_untested=prefill_untested
        )
    
    def validate_file(
        self,
        file_path: str,
        project_id: Optional[int] = None
    ) -> TestImportValidationResult:
        """Validate file before import"""
        return self.import_handler.validate_import(file_path, project_id)
    
    def execute_import(
        self,
        validation_result: TestImportValidationResult,
        update_duplicates: bool = False,
        on_progress: Callable[[str], None] = None
    ) -> TestImportResult:
        """Execute the import with transaction"""
        return self.import_handler.import_with_transaction(
            validation_result,
            update_duplicates,
            on_progress
        )


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def generate_test_template(
    db_manager,
    project_id: Optional[int] = None,
    project_name: str = "",
    output_path: Optional[str] = None
) -> Tuple[Any, str]:
    """Generate a test results template"""
    generator = TestResultsTemplateGenerator(db_manager)
    return generator.generate_template(project_id, project_name, output_path)


def import_test_results(
    db_manager,
    file_path: str,
    project_id: Optional[int] = None,
    update_duplicates: bool = False
) -> Tuple[TestImportValidationResult, TestImportResult]:
    """
    Convenience function to validate and import test results.
    
    Returns:
        Tuple of (validation_result, import_result)
    """
    handler = TestResultsImportHandler(db_manager)
    
    validation = handler.validate_import(file_path, project_id)
    
    if validation.can_import:
        result = handler.import_with_transaction(validation, update_duplicates)
    else:
        result = TestImportResult(
            success=False,
            message="Doğrulama başarısız - içe aktarılacak veri yok"
        )
    
    return validation, result
