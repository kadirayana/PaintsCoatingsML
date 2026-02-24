"""
Paint Formulation AI - Intelligent Excel Workflow
=================================================
Smart template generation with embedded metadata and transactional database integrity.

Features:
- Hidden metadata sheet for project context
- Robust import with project validation ("handshake")
- Atomic database transactions with rollback safety
- Dynamic filename generation
"""

import os
import uuid
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional, Callable, Any
from dataclasses import dataclass, field
from contextlib import contextmanager

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.protection import SheetProtection
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class TemplateMetadata:
    """Metadata embedded in Excel templates"""
    project_id: Optional[int] = None
    project_name: str = ""
    template_version: str = "v2.1"
    generated_timestamp: str = ""
    generated_by: str = "PaintFormulationAI"
    
    def __post_init__(self):
        if not self.generated_timestamp:
            self.generated_timestamp = datetime.now().isoformat()
    
    def to_dict(self) -> Dict:
        return {
            'Project_ID': self.project_id,
            'Project_Name': self.project_name,
            'Template_Version': self.template_version,
            'Generated_Timestamp': self.generated_timestamp,
            'Generated_By': self.generated_by
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'TemplateMetadata':
        return cls(
            project_id=data.get('Project_ID'),
            project_name=data.get('Project_Name', ''),
            template_version=data.get('Template_Version', 'v1.0'),
            generated_timestamp=data.get('Generated_Timestamp', ''),
            generated_by=data.get('Generated_By', '')
        )


@dataclass
class ImportContext:
    """Context for import validation"""
    file_project_id: Optional[int] = None
    file_project_name: str = ""
    current_project_id: Optional[int] = None
    current_project_name: str = ""
    filename: str = ""
    suggested_formula_name: str = ""
    template_version: str = ""
    requires_confirmation: bool = False
    mismatch_warning: str = ""


@dataclass 
class TransactionResult:
    """Result of a transactional import"""
    success: bool
    message: str
    formulation_id: Optional[int] = None
    formulation_code: str = ""
    components_count: int = 0
    rollback_performed: bool = False
    validation_errors: List[str] = field(default_factory=list)


# ============================================================================
# SMART TEMPLATE GENERATOR
# ============================================================================

class SmartTemplateGenerator:
    """
    Generates Excel templates with embedded metadata for intelligent import.
    
    Features:
    - Hidden _SYSTEM_METADATA sheet
    - Pre-filled headers with instructions
    - Dynamic filename generation
    """
    
    METADATA_SHEET_NAME = "_SYSTEM_METADATA"
    DATA_SHEET_NAME = "Formulation_Data"
    INSTRUCTIONS_SHEET_NAME = "Instructions"
    
    # Template columns
    COLUMNS = [
        ('material_code', 'hammadde Kodu', 'Zorunlu'),
        ('material_name', 'hammadde Adı', 'Opsiyonel'),
        ('quantity', 'Miktar (kg)', 'Zorunlu'),
        ('percentage', 'Yüzde (%)', 'Opsiyonel'),
        ('category', 'Kategori', 'Opsiyonel'),
        ('unit_price', 'Birim Fiyat', 'Opsiyonel'),
        ('notes', 'Notlar', 'Opsiyonel')
    ]
    
    def __init__(self, db_manager=None):
        self.db_manager = db_manager
    
    def generate_template(
        self,
        project_id: Optional[int] = None,
        project_name: str = "",
        output_path: Optional[str] = None,
        include_sample_data: bool = True
    ) -> Tuple[str, str]:
        """
        Generate a smart Excel template with embedded metadata.
        
        Args:
            project_id: Project ID to embed
            project_name: Project name
            output_path: Optional output path (if None, returns suggested path)
            include_sample_data: Whether to include sample rows
            
        Returns:
            Tuple of (file_path, suggested_filename)
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl kütüphanesi gerekli. Kurulum: pip install openpyxl")
        
        # Create workbook
        wb = Workbook()
        
        # Create metadata
        metadata = TemplateMetadata(
            project_id=project_id,
            project_name=project_name
        )
        
        # Create sheets
        self._create_data_sheet(wb, include_sample_data)
        self._create_metadata_sheet(wb, metadata)
        self._create_instructions_sheet(wb)
        
        # Generate filename
        suggested_filename = self._generate_filename(project_name)
        
        # Save if path provided
        if output_path:
            if os.path.isdir(output_path):
                file_path = os.path.join(output_path, suggested_filename)
            else:
                file_path = output_path
            
            wb.save(file_path)
            logger.info(f"Template saved: {file_path}")
            return file_path, suggested_filename
        
        return "", suggested_filename
    
    def save_template(self, wb: 'Workbook', file_path: str) -> str:
        """Save workbook to file"""
        wb.save(file_path)
        return file_path
    
    def get_workbook(
        self,
        project_id: Optional[int] = None,
        project_name: str = "",
        include_sample_data: bool = True
    ) -> 'Workbook':
        """Get workbook object without saving (for use with file dialogs)"""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl kütüphanesi gerekli")
        
        wb = Workbook()
        
        metadata = TemplateMetadata(
            project_id=project_id,
            project_name=project_name
        )
        
        self._create_data_sheet(wb, include_sample_data)
        self._create_metadata_sheet(wb, metadata)
        self._create_instructions_sheet(wb)
        
        return wb
    
    def _create_data_sheet(self, wb: 'Workbook', include_sample: bool):
        """Create the main data entry sheet"""
        # Use the default sheet
        ws = wb.active
        ws.title = self.DATA_SHEET_NAME
        
        # Headers
        for col, (key, label, required) in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = cell.font.copy(bold=True)
            
            # Add comment for required fields
            if required == 'Zorunlu':
                cell.value = f"{label} *"
        
        # Column widths
        column_widths = [15, 25, 12, 10, 15, 12, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Sample data
        if include_sample:
            sample_data = [
                ('RESIN-001', 'Epoksi Reçine A', 35, 35, 'binder', 45.0, ''),
                ('PIGMENT-TIO2', 'Titanyum Dioksit', 20, 20, 'pigment', 85.0, ''),
                ('SOLVENT-XYL', 'Ksilen', 25, 25, 'solvent', 12.0, ''),
                ('ADD-001', 'Akış Katkısı', 2, 2, 'additive', 120.0, ''),
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_metadata_sheet(self, wb: 'Workbook', metadata: TemplateMetadata):
        """Create hidden metadata sheet"""
        ws = wb.create_sheet(self.METADATA_SHEET_NAME)
        
        # Write metadata
        meta_dict = metadata.to_dict()
        for row, (key, value) in enumerate(meta_dict.items(), 1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value) if value is not None else '')
        
        # Hide the sheet
        ws.sheet_state = 'hidden'
    
    def _create_instructions_sheet(self, wb: 'Workbook'):
        """Create instructions sheet"""
        ws = wb.create_sheet(self.INSTRUCTIONS_SHEET_NAME)
        
        instructions = [
            "FORMÜLASYON ŞABLONU KULLANIM KILAVUZU",
            "",
            "1. 'Formulation_Data' sayfasına hammaddelerinizi girin",
            "2. Zorunlu alanlar (*) doldurulmalıdır:",
            "   - hammadde Kodu: Benzersiz hammadde tanımlayıcısı",
            "   - Miktar: kg cinsinden miktar",
            "",
            "3. Opsiyonel alanlar:",
            "   - hammadde Adı: Otomatik tanınmayan hammaddeler için",
            "   - Yüzde: Otomatik hesaplanır (boş bırakılabilir)",
            "   - Kategori: binder, pigment, solvent, additive",
            "",
            "4. Dosyayı kaydedin ve uygulamaya yükleyin",
            "",
            "NOT: '_SYSTEM_METADATA' sayfasını DEĞİŞTİRMEYİN!",
            "Bu sayfa proje bilgilerini içerir."
        ]
        
        for row, text in enumerate(instructions, 1):
            ws.cell(row=row, column=1, value=text)
        
        ws.column_dimensions['A'].width = 60
    
    def _generate_filename(self, project_name: str) -> str:
        """Generate standardized filename"""
        date_str = datetime.now().strftime("%Y%m%d")
        
        # Sanitize project name
        safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_'))
        safe_name = safe_name.strip().replace(' ', '_') or "New"
        
        return f"Formulation_{safe_name}_{date_str}.xlsx"


# ============================================================================
# INTELLIGENT IMPORT HANDLER
# ============================================================================

class IntelligentImportHandler:
    """
    Handles Excel import with metadata validation and transactional integrity.
    
    Features:
    - Reads embedded metadata for project context
    - Validates project match (the "handshake")
    - Atomic database transactions
    - Comprehensive error handling
    """
    
    METADATA_SHEET_NAME = "_SYSTEM_METADATA"
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
    
    def read_import_context(
        self,
        file_path: str,
        current_project_id: Optional[int] = None,
        current_project_name: str = ""
    ) -> ImportContext:
        """
        Read metadata and prepare import context (the "handshake").
        
        This should be called BEFORE the actual import to validate context.
        
        Args:
            file_path: Path to Excel file
            current_project_id: Currently open project ID in UI
            current_project_name: Currently open project name
            
        Returns:
            ImportContext with validation status
        """
        context = ImportContext(
            current_project_id=current_project_id,
            current_project_name=current_project_name,
            filename=os.path.basename(file_path)
        )
        
        # Generate suggested formula name from filename
        base_name = os.path.splitext(context.filename)[0]
        context.suggested_formula_name = base_name
        
        try:
            if not HAS_OPENPYXL:
                logger.warning("openpyxl not available, skipping metadata read")
                return context
            
            wb = load_workbook(file_path, read_only=True)
            
            # Read metadata if exists
            if self.METADATA_SHEET_NAME in wb.sheetnames:
                metadata = self._read_metadata_sheet(wb[self.METADATA_SHEET_NAME])
                
                context.file_project_id = metadata.project_id
                context.file_project_name = metadata.project_name
                context.template_version = metadata.template_version
                
                # Check for project mismatch
                if metadata.project_id and current_project_id:
                    if metadata.project_id != current_project_id:
                        context.requires_confirmation = True
                        context.mismatch_warning = (
                            f"Bu dosya '{metadata.project_name}' projesine ait.\n"
                            f"Şu anda '{current_project_name}' projesi açık.\n\n"
                            f"Yine de mevcut projeye aktarmak istiyor musunuz?"
                        )
            
            wb.close()
            
        except Exception as e:
            logger.warning(f"Could not read metadata: {e}")
        
        return context
    
    def _read_metadata_sheet(self, ws) -> TemplateMetadata:
        """Read metadata from hidden sheet"""
        data = {}
        
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                key = str(row[0])
                value = row[1]
                
                # Parse Project_ID as int
                if key == 'Project_ID' and value:
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        value = None
                
                data[key] = value
        
        return TemplateMetadata.from_dict(data)
    
    def import_with_transaction(
        self,
        file_path: str,
        project_id: Optional[int],
        formula_name: str = "",
        formula_code: Optional[str] = None,
        on_progress: Callable[[str], None] = None
    ) -> TransactionResult:
        """
        Import Excel file with atomic database transaction.
        
        Implements the full transaction flow:
        1. Begin transaction
        2. Validate all materials
        3. Create formulation header
        4. Bulk insert components
        5. Commit or Rollback
        
        Args:
            file_path: Path to Excel file
            project_id: Target project ID
            formula_name: Name for the formulation
            formula_code: Optional code (auto-generated if not provided)
            on_progress: Progress callback
            
        Returns:
            TransactionResult
        """
        def log_progress(msg: str):
            logger.info(msg)
            if on_progress:
                on_progress(msg)
        
        if not HAS_PANDAS:
            return TransactionResult(
                success=False,
                message="Pandas kütüphanesi gerekli"
            )
        
        log_progress("📖 Excel dosyası okunuyor...")
        
        try:
            # Read Excel data
            df = pd.read_excel(file_path, sheet_name=0)
            
            if df.empty:
                return TransactionResult(
                    success=False,
                    message="Excel dosyası boş"
                )
            
            # Normalize columns
            df = self._normalize_columns(df)
            
            # Generate formula code if not provided
            if not formula_code:
                formula_code = self._generate_formula_code()
            
            if not formula_name:
                formula_name = os.path.splitext(os.path.basename(file_path))[0]
            
            log_progress("🔍 hammaddeler doğrulanıyor...")
            
            # ================================================================
            # START TRANSACTION
            # ================================================================
            with self._transaction_context() as tx:
                try:
                    # Step A: Validate all materials
                    validation_errors = []
                    components_data = []
                    
                    for idx, row in df.iterrows():
                        material_code = str(row.get('material_code', '')).strip()
                        
                        if not material_code:
                            continue
                        
                        # Look up material
                        material = self._lookup_or_create_material(
                            tx['cursor'],
                            code=material_code,
                            name=str(row.get('material_name', material_code)).strip(),
                            category=str(row.get('category', 'other')).strip()
                        )
                        
                        if material is None:
                            validation_errors.append(f"Satır {idx+2}: '{material_code}' bulunamadı")
                            continue
                        
                        quantity = self._parse_number(row.get('quantity', 0))
                        percentage = self._parse_number(row.get('percentage', 0))
                        
                        components_data.append({
                            'material_id': material['id'],
                            'material_name': material['name'],
                            'material_code': material_code,
                            'amount': quantity,
                            'percentage': percentage
                        })
                    
                    if not components_data:
                        raise ValueError("Geçerli bileşen bulunamadı")
                    
                    log_progress(f"✅ {len(components_data)} hammadde doğrulandı")
                    
                    # Step B: Create formulation header
                    log_progress("📝 Formülasyon kaydı oluşturuluyor...")
                    
                    tx['cursor'].execute('''
                        INSERT INTO formulations (project_id, formula_code, formula_name, status)
                        VALUES (?, ?, ?, 'imported')
                    ''', (project_id, formula_code, formula_name))
                    
                    formulation_id = tx['cursor'].lastrowid
                    
                    # Step C: Bulk insert components
                    log_progress("📦 Bileşenler ekleniyor...")
                    
                    for comp in components_data:
                        # Insert into components table
                        tx['cursor'].execute('''
                            INSERT INTO components (formulation_id, component_name, component_type, amount, percentage)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (
                            formulation_id,
                            comp['material_name'],
                            comp['material_code'],
                            comp['amount'],
                            comp['percentage']
                        ))
                        
                        # Insert into formulation_materials table
                        tx['cursor'].execute('''
                            INSERT INTO formulation_materials (formulation_id, material_id, amount)
                            VALUES (?, ?, ?)
                        ''', (formulation_id, comp['material_id'], comp['amount']))
                    
                    # COMMIT
                    tx['connection'].commit()
                    log_progress("✅ İçe aktarım başarılı!")
                    
                    return TransactionResult(
                        success=True,
                        message=f"Formülasyon '{formula_name}' başarıyla oluşturuldu",
                        formulation_id=formulation_id,
                        formulation_code=formula_code,
                        components_count=len(components_data),
                        validation_errors=validation_errors
                    )
                    
                except Exception as e:
                    # ROLLBACK
                    tx['connection'].rollback()
                    logger.error(f"Transaction rollback: {e}")
                    
                    return TransactionResult(
                        success=False,
                        message=f"İçe aktarım hatası: {str(e)}",
                        rollback_performed=True,
                        validation_errors=validation_errors if 'validation_errors' in dir() else []
                    )
                    
        except Exception as e:
            logger.error(f"Import error: {e}")
            return TransactionResult(
                success=False,
                message=f"Dosya okuma hatası: {str(e)}"
            )
    
    @contextmanager
    def _transaction_context(self):
        """Context manager for database transaction"""
        conn = self.db_manager.get_connection().__enter__()
        cursor = conn.cursor()
        
        try:
            yield {'connection': conn, 'cursor': cursor}
        finally:
            conn.close()
    
    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize column names"""
        column_map = {}
        
        mappings = {
            'material_code': ['hammadde kodu', 'code', 'kod', 'material code', 'material_code'],
            'material_name': ['hammadde adı', 'hammadde adi', 'name', 'ad', 'material name'],
            'quantity': ['miktar', 'amount', 'qty', 'quantity'],
            'percentage': ['yüzde', 'yuzde', '%', 'percentage', 'oran'],
            'category': ['kategori', 'category', 'type', 'tip'],
            'unit_price': ['fiyat', 'price', 'birim fiyat', 'unit_price']
        }
        
        for col in df.columns:
            col_lower = str(col).lower().strip()
            for standard, variations in mappings.items():
                if col_lower in variations:
                    column_map[col] = standard
                    break
        
        return df.rename(columns=column_map)
    
    def _lookup_or_create_material(
        self,
        cursor,
        code: str,
        name: str,
        category: str
    ) -> Optional[Dict]:
        """Look up material or create as incomplete"""
        # Try to find by code
        cursor.execute('SELECT * FROM materials WHERE code = ? OR name = ?', (code, code))
        row = cursor.fetchone()
        
        if row:
            return dict(row)
        
        # Create as incomplete
        cursor.execute('''
            INSERT INTO materials (name, code, category, is_incomplete)
            VALUES (?, ?, ?, 1)
        ''', (name or code, code, category))
        
        return {
            'id': cursor.lastrowid,
            'name': name or code,
            'code': code,
            'is_incomplete': True
        }
    
    def _generate_formula_code(self) -> str:
        """Generate unique formula code"""
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        return f"IMP-{timestamp}"
    
    def _parse_number(self, value) -> float:
        """Parse number from value"""
        if pd.isna(value):
            return 0.0
        try:
            if isinstance(value, str):
                value = value.replace(',', '.').replace('%', '').strip()
            return float(value)
        except (ValueError, TypeError):
            return 0.0


# ============================================================================
# INTEGRATED WORKFLOW CONTROLLER
# ============================================================================

class IntelligentExcelWorkflow:
    """
    High-level controller for the intelligent Excel workflow.
    
    Combines template generation, import handling, and UI integration.
    """
    
    def __init__(self, db_manager, on_formulation_created: Callable = None):
        """
        Args:
            db_manager: LocalDBManager instance
            on_formulation_created: Callback when formulation is created (id, code)
        """
        self.db_manager = db_manager
        self.on_formulation_created = on_formulation_created
        
        self.template_generator = SmartTemplateGenerator(db_manager)
        self.import_handler = IntelligentImportHandler(db_manager)
    
    def download_template(
        self,
        project_id: Optional[int] = None,
        project_name: str = ""
    ) -> Tuple[Any, str]:
        """
        Get template workbook for download.
        
        Returns:
            Tuple of (Workbook, suggested_filename)
        """
        wb = self.template_generator.get_workbook(
            project_id=project_id,
            project_name=project_name,
            include_sample_data=True
        )
        filename = self.template_generator._generate_filename(project_name)
        return wb, filename
    
    def prepare_import(
        self,
        file_path: str,
        current_project_id: Optional[int] = None,
        current_project_name: str = ""
    ) -> ImportContext:
        """
        Prepare import by reading context and validating.
        
        Returns:
            ImportContext with validation info
        """
        return self.import_handler.read_import_context(
            file_path,
            current_project_id,
            current_project_name
        )
    
    def execute_import(
        self,
        file_path: str,
        project_id: Optional[int],
        formula_name: str = "",
        formula_code: Optional[str] = None,
        on_progress: Callable[[str], None] = None
    ) -> TransactionResult:
        """
        Execute the full import with transaction.
        
        Returns:
            TransactionResult
        """
        result = self.import_handler.import_with_transaction(
            file_path=file_path,
            project_id=project_id,
            formula_name=formula_name,
            formula_code=formula_code,
            on_progress=on_progress
        )
        
        # Callback on success
        if result.success and self.on_formulation_created:
            self.on_formulation_created(result.formulation_id, result.formulation_code)
        
        return result
    
    def get_incomplete_materials_count(self) -> int:
        """Get count of incomplete materials for notification"""
        try:
            materials = self.db_manager.get_incomplete_materials()
            return len(materials)
        except Exception:
            return 0


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def generate_smart_template(
    db_manager,
    project_id: Optional[int] = None,
    project_name: str = "",
    output_path: Optional[str] = None
) -> Tuple[str, str]:
    """
    Convenience function to generate a smart template.
    
    Returns:
        Tuple of (file_path, suggested_filename)
    """
    generator = SmartTemplateGenerator(db_manager)
    return generator.generate_template(project_id, project_name, output_path)


def import_excel_with_transaction(
    db_manager,
    file_path: str,
    project_id: Optional[int],
    formula_name: str = ""
) -> TransactionResult:
    """
    Convenience function for transactional import.
    
    Returns:
        TransactionResult
    """
    handler = IntelligentImportHandler(db_manager)
    return handler.import_with_transaction(file_path, project_id, formula_name)
